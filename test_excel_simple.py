"""
Simple Excel generation test
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

print("Testing Excel generation...")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Test Sheet"

# Add header
ws['A1'] = "Test Excel File"
ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Add some data
ws['A3'] = "Date:"
ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

ws['A4'] = "Status:"
ws['B4'] = "✅ openpyxl is working!"

ws['A5'] = "Test 1:"
ws['B5'] = "Excel file creation"

ws['A6'] = "Test 2:"
ws['B6'] = "Styling and formatting"

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 30

# Save file
output_file = 'test_excel_output.xlsx'
wb.save(output_file)

print(f"✅ Success! Excel file created: {output_file}")
print(f"   File location: c:/Users/Silvi/Projects/trading-bot/{output_file}")
