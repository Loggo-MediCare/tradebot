"""
批量运行所有台股的交易信号生成器 (输出到Excel)
================================
自动运行所有已训练的台股模型的交易信号
每个股票的输出保存到单独的工作表

最新添加的XGBoost模型股票 (2026-03-02):
==========================================
Stock      | Name           | Accuracy | Model Type | Status
-----------|----------------|----------|------------|--------
3563.TW    | 牧德           | 60.28%   | XGBoost    | ✅
3576.TW    | 聯合再生       | 68.15%   | XGBoost    | ✅
3615.TWO   | 安可           | 67.54%   | XGBoost    | ✅
3665.TW    | 貿聯-KY        | 52.82%   | XGBoost    | ✅
4564.TW    | 元翎           | 65.32%   | XGBoost    | ✅
4577.TWO   | 達航科技       | 51.42%   | XGBoost    | ✅
4768.TWO   | 晶呈科技       | 50.97%   | XGBoost    | ✅
4989.TW    | 榮科           | 64.24%   | XGBoost    | ✅
4991.TWO   | 環宇-KY        | 53.83%   | XGBoost    | ✅
6220.TWO   | 岳豐           | 75.60%   | XGBoost    | 🌟 EXCELLENT
6230.TW    | 尼得科超眾     | 65.93%   | XGBoost    | ✅
6442.TW    | 光聖           | 50.00%   | XGBoost    | ✅
6526.TW    | 達發           | 49.28%   | XGBoost    | ⚠️
6789.TW    | 采鈺           | 56.77%   | XGBoost    | ✅
6830.TW    | 汎銓           | 62.50%   | XGBoost    | ✅
6877.TWO   | 鏵友益         | 70.75%   | XGBoost    | 🌟 EXCELLENT
8438.TW    | 昶昕           | 53.97%   | XGBoost    | ✅
8927.TWO   | 北基           | 67.74%   | XGBoost    | ✅

Top Performers (≥70%):
- 6220.TWO 岳豐: 75.60%
- 6877.TWO 鏵友益: 70.75%

All models trained: 2026-03-02
All signal files created: 2026-03-02
Added to run_all: 2026-03-02
"""

import subprocess
import sys
import io
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import yfinance as yf
import numpy as np
import pandas as pd

# 修复 Windows 编码问题
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 强制输出刷新
os.environ['PYTHONIOENCODING'] = 'utf-8'
os.environ['PYTHONUNBUFFERED'] = '1'

# 获取脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 台股信号生成器 (用戶指定順序)
SIGNAL_SCRIPTS = [
    {'file': 'get_trading_signal_2330.py', 'name': '2330 台積電'},
    {'file': 'get_trading_signal_2317.py', 'name': '2317 鴻海'},
    {'file': 'get_trading_signal_6515.py', 'name': '6515 穎崴'},
    {'file': 'get_trading_signal_2408.py', 'name': '2408 南亞科'},
    {'file': 'get_trading_signal_2308.py', 'name': '2308 台達電'},
    {'file': 'get_trading_signal_2313.py', 'name': '2313 華通'},
    {'file': 'get_trading_signal_2454.py', 'name': '2454 聯發科'},
    {'file': 'get_trading_signal_2485.py', 'name': '2485 兆赫'},
    {'file': 'get_trading_signal_2337.py', 'name': '2337 旺宏'},
    {'file': 'get_trading_signal_2344.py', 'name': '2344 華邦電'},
    {'file': 'get_trading_signal_2367.py', 'name': '2367 燿華'},
    {'file': 'get_trading_signal_3481.py', 'name': '3481 群創'},
    {'file': 'get_trading_signal_2603.py', 'name': '2603 長榮'},
    {'file': 'get_trading_signal_6770.py', 'name': '6770 力積電'},
    {'file': 'get_trading_signal_3665.py', 'name': '3665 貿聯-KY'},
    {'file': 'get_trading_signal_3017.py', 'name': '3017 奇鋐'},
    {'file': 'get_trading_signal_3711.py', 'name': '3711 日月光投控'},
    {'file': 'get_trading_signal_3037.py', 'name': '3037 欣興'},
    {'file': 'get_trading_signal_2327.py', 'name': '2327 國巨'},
    {'file': 'get_trading_signal_2382.py', 'name': '2382 廣達'},
    {'file': 'get_trading_signal_3443.py', 'name': '3443 創意'},
    {'file': 'get_trading_signal_2383.py', 'name': '2383 台光電'},
    {'file': 'get_trading_signal_6442.py', 'name': '6442 光聖'},
    {'file': 'get_trading_signal_3661.py', 'name': '3661 世芯-KY'},
    {'file': 'get_trading_signal_6669.py', 'name': '6669 緯穎'},
    {'file': 'get_trading_signal_6683.py', 'name': '6683 雍智科技'},
    {'file': 'get_trading_signal_3231.py', 'name': '3231 緯創'},
    {'file': 'get_trading_signal_2303.py', 'name': '2303 聯電'},
    {'file': 'get_trading_signal_2368.py', 'name': '2368 金像電'},
    {'file': 'get_trading_signal_2345.py', 'name': '2345 智邦'},
    {'file': 'get_trading_signal_1303.py', 'name': '1303 南亞'},
    {'file': 'get_trading_signal_2360.py', 'name': '2360 致茂'},
    {'file': 'get_trading_signal_2449.py', 'name': '2449 京元電子'},
    {'file': 'get_trading_signal_6443.py', 'name': '6443 元晶'},
    {'file': 'get_trading_signal_4989.py', 'name': '4989 榮科'},
    {'file': 'get_trading_signal_6285.py', 'name': '6285 啟碁'},
    {'file': 'get_trading_signal_3715.py', 'name': '3715 定穎投控'},
    {'file': 'get_trading_signal_3563.py', 'name': '3563 牧德'},
    {'file': 'get_trading_signal_3653.py', 'name': '3653 健策'},
    {'file': 'get_trading_signal_2891.py', 'name': '2891 中信金'},
    {'file': 'get_trading_signal_6239.py', 'name': '6239 力成'},
    {'file': 'get_trading_signal_3533.py', 'name': '3533 嘉澤'},
    {'file': 'get_trading_signal_8069.py', 'name': '8069 元太'},
    {'file': 'get_trading_signal_6223.py', 'name': '6223 旺矽'},
    {'file': 'get_trading_signal_3363.py', 'name': '3363 上詮'},
    {'file': 'get_trading_signal_3449.py', 'name': '3449 鈺德'},
    {'file': 'get_trading_signal_5483.py', 'name': '5483 中美晶'},
    {'file': 'get_trading_signal_6163.py', 'name': '6163 華電網'},
    {'file': 'get_trading_signal_7709.py', 'name': '7709 榮田'},
    {'file': 'get_trading_signal_7717.py', 'name': '7717 萊德光電'},
    {'file': 'get_trading_signal_3260.py', 'name': '3260 威剛'},
    {'file': 'get_trading_signal_3491.py', 'name': '3491 昇達科'},
    {'file': 'get_trading_signal_5371.py', 'name': '5371 中光電'},
    {'file': 'get_trading_signal_3105.py', 'name': '3105 穩懋'},
    {'file': 'get_trading_signal_4971.py', 'name': '4971 IET-KY'},
    {'file': 'get_trading_signal_6187.py', 'name': '6187 萬潤'},
    {'file': 'get_trading_signal_3615.py', 'name': '3615 安可'},
    {'file': 'get_trading_signal_4577.py', 'name': '4577 達航科技'},
    {'file': 'get_trading_signal_4768.py', 'name': '4768 晶呈科技'},
    {'file': 'get_trading_signal_4991.py', 'name': '4991 環宇-KY'},
    {'file': 'get_trading_signal_6220.py', 'name': '6220 岳豐'},
    {'file': 'get_trading_signal_6877.py', 'name': '6877 鏵友益'},
    {'file': 'get_trading_signal_8927.py', 'name': '8927 北基'},
    {'file': 'get_trading_signal_1519.py', 'name': '1519 華城'},
    {'file': 'get_trading_signal_6805.py', 'name': '6805 富世達'},
    {'file': 'get_trading_signal_6789.py', 'name': '6789 采鈺'},
    {'file': 'get_trading_signal_8021.py', 'name': '8021 尖點'},
    {'file': 'get_trading_signal_3006.py', 'name': '3006 晶豪科'},
    {'file': 'get_trading_signal_6830.py', 'name': '6830 汎銓'},
    {'file': 'get_trading_signal_2357.py', 'name': '2357 華碩'},
    {'file': 'get_trading_signal_3030.py', 'name': '3030 德律'},
    {'file': 'get_trading_signal_2409.py', 'name': '2409 友達'},
    {'file': 'get_trading_signal_2376.py', 'name': '2376 技嘉'},
    {'file': 'get_trading_signal_8210.py', 'name': '8210 勤誠'},
    {'file': 'get_trading_signal_6446.py', 'name': '6446 藥華藥'},
    {'file': 'get_trading_signal_1326.py', 'name': '1326 台塑化'},
    {'file': 'get_trading_signal_8046.py', 'name': '8046 南電'},
    {'file': 'get_trading_signal_1605.py', 'name': '1605 華新'},
    {'file': 'get_trading_signal_1301.py', 'name': '1301 台塑'},
    {'file': 'get_trading_signal_2059.py', 'name': '2059 川湖'},
    {'file': 'get_trading_signal_6781.py', 'name': '6781 AES-KY'},
    {'file': 'get_trading_signal_2884.py', 'name': '2884 玉山金'},
    {'file': 'get_trading_signal_6271.py', 'name': '6271 同欣電'},
    {'file': 'get_trading_signal_2002.py', 'name': '2002 中鋼'},
    {'file': 'get_trading_signal_6526.py', 'name': '6526 達發'},
    {'file': 'get_trading_signal_3138.py', 'name': '3138 耀登'},
    {'file': 'get_trading_signal_8150.py', 'name': '8150 南茂'},
    {'file': 'get_trading_signal_1101.py', 'name': '1101 台泥'},
    {'file': 'get_trading_signal_2890.py', 'name': '2890 永豐金'},
    {'file': 'get_trading_signal_3044.py', 'name': '3044 健鼎'},
    {'file': 'get_trading_signal_4967.py', 'name': '4967 十銓'},
    {'file': 'get_trading_signal_2451.py', 'name': '2451 創見資訊'},
    {'file': 'get_trading_signal_8110.py', 'name': '8110 華東'},
    {'file': 'get_trading_signal_2385.py', 'name': '2385 群光'},
    {'file': 'get_trading_signal_4938.py', 'name': '4938 和碩'},
    {'file': 'get_trading_signal_3576.py', 'name': '3576 聯合再生'},
    {'file': 'get_trading_signal_2634.py', 'name': '2634 漢翔'},
    {'file': 'get_trading_signal_1514.py', 'name': '1514 亞力'},
    {'file': 'get_trading_signal_4722.py', 'name': '4722 国精化'},
    {'file': 'get_trading_signal_6472.py', 'name': '6472 保瑞'},
    {'file': 'get_trading_signal_8131.py', 'name': '8131 福懋科'},
    {'file': 'get_trading_signal_6230.py', 'name': '6230 尼得科超眾'},
    {'file': 'get_trading_signal_2363.py', 'name': '2363 矽統'},
    {'file': 'get_trading_signal_6209.py', 'name': '6209 今國光'},
    {'file': 'get_trading_signal_3135.py', 'name': '3135 凌航'},
    {'file': 'get_trading_signal_6269.py', 'name': '6269 台郡'},
    {'file': 'get_trading_signal_8438.py', 'name': '8438 昶昕'},
    {'file': 'get_trading_signal_4564.py', 'name': '4564 元翎'},
    {'file': 'get_trading_signal_4540.py', 'name': '4540 全球傳動'},
    {'file': 'get_trading_signal_8499.py', 'name': '8499 鼎炫-KY'},
    {'file': 'get_trading_signal_6477.py', 'name': '6477 安集'},
    {'file': 'get_trading_signal_3004.py', 'name': '3004 豐達科'},
    {'file': 'get_trading_signal_4746.py', 'name': '4746 台耀'},
    {'file': 'get_trading_signal_8222.py', 'name': '8222 寶一'},
    {'file': 'get_trading_signal_3022.py', 'name': '3022 威強電'},
    {'file': 'get_trading_signal_6668.py', 'name': '6668 中揚光'},
    {'file': 'get_trading_signal_2314.py', 'name': '2314 台揚'},
    {'file': 'get_trading_signal_1314.py', 'name': '1314 中石化'},
    {'file': 'get_trading_signal_8908.py', 'name': '8908 欣雄'},
    {'file': 'get_trading_signal_9931.py', 'name': '9931 欣高'},
    {'file': 'get_trading_signal_8917.py', 'name': '8917 欣泰'},
    {'file': 'get_trading_signal_6505.py', 'name': '6505 台塑化'},
    {'file': 'get_trading_signal_9918.py', 'name': '9918 欣天然'},
    {'file': 'get_trading_signal_2412.py', 'name': '2412 中華電信'},
    {'file': 'get_trading_signal_6274.py', 'name': '6274 台燿'},
    {'file': 'get_trading_signal_8112.py', 'name': '8112 至上'},
    {'file': 'get_trading_signal_2049.py', 'name': '2049 上銀'},
    {'file': 'get_trading_signal_1785.py', 'name': '1785 光洋科'},
    {'file': 'get_trading_signal_6531.py', 'name': '6531 愛普'},
    {'file': 'get_trading_signal_2395.py', 'name': '2395 研華'},
    {'file': 'get_trading_signal_4749.py', 'name': '4749 不二家'},
    {'file': 'get_trading_signal_3131.py', 'name': '3131 弘塑'},
    {'file': 'get_trading_signal_3081.py', 'name': '3081 聯亞科技'},
    {'file': 'get_trading_signal_6510.py', 'name': '6510 精測電子'},
    {'file': 'get_trading_signal_3535.py', 'name': '3535 晶彩科'},
    {'file': 'get_trading_signal_8064.py', 'name': '8064 東捷科技'},
    {'file': 'get_trading_signal_3163.py', 'name': '3163 Browave'},
    {'file': 'get_trading_signal_3455.py', 'name': '3455 由田'},
    {'file': 'get_trading_signal_2426.py', 'name': '2426 鼎元'},
    {'file': 'get_trading_signal_3583.py', 'name': '3583 辛耘'},
    {'file': 'get_trading_signal_8028.py', 'name': '8028 昇陽半導體'},
    {'file': 'get_trading_signal_3680.py', 'name': '3680 家登'},
    {'file': 'get_trading_signal_4772.py', 'name': '4772 台特化'},
    {'file': 'get_trading_signal_6788.py', 'name': '6788 華景電'},
    {'file': 'get_trading_signal_7703.py', 'name': '7703 銳澤'},
    {'file': 'get_trading_signal_8147.py', 'name': '8147 大綜'},
    {'file': 'get_trading_signal_2404.py', 'name': '2404 漢唐'},
    {'file': 'get_trading_signal_6196.py', 'name': '6196 帆宣'},
    {'file': 'get_trading_signal_6605.py', 'name': '6605 信紘科'},
    {'file': 'get_trading_signal_6139.py', 'name': '6139 亞翔'},
    {'file': 'get_trading_signal_8071.py', 'name': '8071 竹陸科技'},
    {'file': 'get_trading_signal_1560.py', 'name': '1560 中砂'},
    {'file': 'get_trading_signal_6438.py', 'name': '6438 迅得'},
    {'file': 'get_trading_signal_6449.py', 'name': '6449 倍利科'},
    {'file': 'get_trading_signal_8027.py', 'name': '8027 鈦昇'},
    {'file': 'get_trading_signal_5351.py', 'name': '5351 鈺創'},
    {'file': 'get_trading_signal_4720.py', 'name': '4720 德淵'},
    {'file': 'get_trading_signal_6176.py', 'name': '6176 瑞儀'},
    {'file': 'get_trading_signal_3380.py', 'name': '3380 明泰'},
    {'file': 'get_trading_signal_6672.py', 'name': '6672 騰輝電子KY'},
    {'file': 'get_trading_signal_6213.py', 'name': '6213 聯茂'},
    {'file': 'get_trading_signal_7734.py', 'name': '7734 印能科技'},
    {'file': 'get_trading_signal_7751.py', 'name': '7751 竑騰'},
    {'file': 'get_trading_signal_2486.py', 'name': '2486 一詮'},
    {'file': 'get_trading_signal_6138.py', 'name': '6138 茂達'},
    {'file': 'get_trading_signal_8103.py', 'name': '8103 瀚荃'},
    {'file': 'get_trading_signal_1569.py', 'name': '1569 濱川'},
    {'file': 'get_trading_signal_1595.py', 'name': '1595 川寶'},
    {'file': 'get_trading_signal_6108.py', 'name': '6108 競國'},
    {'file': 'get_trading_signal_4951.py', 'name': '4951 精拓科'},
    {'file': 'get_trading_signal_1727.py', 'name': '1727 中華化'},
    {'file': 'get_trading_signal_6234.py', 'name': '6234 高僑'},
    {'file': 'get_trading_signal_6488.py', 'name': '6488 環球晶'},
    {'file': 'get_trading_signal_6207.py', 'name': '6207 雷科'},
    {'file': 'get_trading_signal_6937.py', 'name': '6937 天虹'},
    {'file': 'get_trading_signal_3189.py', 'name': '3189 景碩'},
    {'file': 'get_trading_signal_6147.py', 'name': '6147 頎邦'},
    {'file': 'get_trading_signal_3624.py', 'name': '3624 光頡'},
    {'file': 'get_trading_signal_8455.py', 'name': '8455 大拓-KY'},
    {'file': 'get_trading_signal_6924.py', 'name': '6924 榮惠-KY'},
    {'file': 'get_trading_signal_3577.py', 'name': '3577 泓格'},
    {'file': 'get_trading_signal_8374.py', 'name': '8374 羅昇'},
    {'file': 'get_trading_signal_2359.py', 'name': '2359 所羅門'},
    {'file': 'get_trading_signal_3236.py', 'name': '3236 千如'},
    {'file': 'get_trading_signal_6204.py', 'name': '6204 艾華'},
    {'file': 'get_trading_signal_3024.py', 'name': '3024 憶聲'},
    {'file': 'get_trading_signal_6432.py', 'name': '6432 今展科'},
    {'file': 'get_trading_signal_3609.py', 'name': '3609 三一東林'},
    {'file': 'get_trading_signal_8299.py', 'name': '8299 群聯'},
    {'file': 'get_trading_signal_3581.py', 'name': '3581 博磊'},
    {'file': 'get_trading_signal_8291.py', 'name': '8291 尚茂'},
    {'file': 'get_trading_signal_3265.py', 'name': '3265 台星科'},
    {'file': 'get_trading_signal_3714.py', 'name': '3714 富采'},
    {'file': 'get_trading_signal_2340.py', 'name': '2340 台亞'},
    {'file': 'get_trading_signal_1773.py', 'name': '1773 勝一'},
    {'file': 'get_trading_signal_5215.py', 'name': '5215 科嘉-KY'},
    {'file': 'get_trading_signal_3587.py', 'name': '3587 閎康'},
    {'file': 'get_trading_signal_3691.py', 'name': '3691 碩禾'},
    {'file': 'get_trading_signal_3264.py', 'name': '3264 欣銓'},
    {'file': 'get_trading_signal_6257.py', 'name': '6257 矽格'},
    {'file': 'get_trading_signal_3055.py', 'name': '3055 蔚華科'},
    {'file': 'get_trading_signal_5289.py', 'name': '5289 宜鼎'},
    {'file': 'get_trading_signal_7610.py', 'name': '7610 聯友金屬'},
    {'file': 'get_trading_signal_7788.py', 'name': '7788 兆川精密'},
    {'file': 'get_trading_signal_2481.py', 'name': '2481 良風'},
    {'file': 'get_trading_signal_3023.py', 'name': '3023 信昌電'},
    {'file': 'get_trading_signal_3663.py', 'name': '3663 健策'},
    {'file': 'get_trading_signal_6538.py', 'name': '6538 倍力'},
    {'file': 'get_trading_signal_3580.py', 'name': '3580 台亞半導體'},
    {'file': 'get_trading_signal_2355.py', 'name': '2355 英業達'},
    {'file': 'get_trading_signal_8044.py', 'name': '8044 網路通'},
    {'file': 'get_trading_signal_3147.py', 'name': '3147 大晶-KY'},
    {'file': 'get_trading_signal_6980.py', 'name': '6980 鐳洋'},
    {'file': 'get_trading_signal_2428.py', 'name': '2428 興勤'},
    {'file': 'get_trading_signal_1597.py', 'name': '1597 直得'},
    {'file': 'get_trading_signal_2455.py', 'name': '2455 全新'},
    {'file': 'get_trading_signal_3026.py', 'name': '3026 禾伸堂'},
    {'file': 'get_trading_signal_2851.py', 'name': '2851 中再保'},
    {'file': 'get_trading_signal_6146.py', 'name': '6146 耕興'},
    {'file': 'get_trading_signal_6706.py', 'name': '6706 惠特'},
    {'file': 'get_trading_signal_3152.py', 'name': '3152 環德'},
    {'file': 'get_trading_signal_6643.py', 'name': '6643 M31'},
    {'file': 'get_trading_signal_6533.py', 'name': '6533 晶心科'},
    {'file': 'get_trading_signal_1717.py', 'name': '1717 長興'},
    {'file': 'get_trading_signal_3008.py', 'name': '3008 大立光'},
    {'file': 'get_trading_signal_3028.py', 'name': '3028 增你強'},
    {'file': 'get_trading_signal_1795.py', 'name': '1795 美時'},
    {'file': 'get_trading_signal_1815.py', 'name': '1815 富喬工業'},
    {'file': 'get_trading_signal_2379.py', 'name': '2379 瑞昱'},
    {'file': 'get_trading_signal_3013.py', 'name': '3013 晟鈦'},
    {'file': 'get_trading_signal_3034.py', 'name': '3034 聯詠'},
    {'file': 'get_trading_signal_3413.py', 'name': '3413 京鼎'},
    {'file': 'get_trading_signal_3450.py', 'name': '3450 聯鈞'},
    {'file': 'get_trading_signal_4958.py', 'name': '4958 臻鼎-KY'},
    {'file': 'get_trading_signal_6640.py', 'name': '6640 嘉聯益'},
    {'file': 'get_trading_signal_6664.py', 'name': '6664 群聯工業'},
    {'file': 'get_trading_signal_7769.py', 'name': '7769 霖揚'},
    {'file': 'get_trading_signal_8996.py', 'name': '8996 高力'},
    {'file': 'get_trading_signal_01810.py', 'name': '01810 小米集團'},
    {'file': 'get_trading_signal_02202.py', 'name': '02202 萬科企業'},
    {'file': 'get_trading_signal_1425.py', 'name': '1425 宇隆'},
    {'file': 'get_trading_signal_1471.py', 'name': '1471 首利'},
    {'file': 'get_trading_signal_1504.py', 'name': '1504 東元'},
    {'file': 'get_trading_signal_1513.py', 'name': '1513 中興電'},
    {'file': 'get_trading_signal_1529.py', 'name': '1529 樂士'},
    {'file': 'get_trading_signal_1582.py', 'name': '1582 信錦'},
    {'file': 'get_trading_signal_1711.py', 'name': '1711 永光'},
    {'file': 'get_trading_signal_1802.py', 'name': '1802 台玻'},
    {'file': 'get_trading_signal_2301.py', 'name': '2301 光寶科'},
    {'file': 'get_trading_signal_2324.py', 'name': '2324 仁寶'},
    {'file': 'get_trading_signal_2331.py', 'name': '2331 精英'},
    {'file': 'get_trading_signal_2356.py', 'name': '2356 英業達'},
    {'file': 'get_trading_signal_2369.py', 'name': '2369 菱生'},
    {'file': 'get_trading_signal_2375.py', 'name': '2375 凱美'},
    {'file': 'get_trading_signal_2377.py', 'name': '2377 微星科技'},
    {'file': 'get_trading_signal_2380.py', 'name': '2380 虹光'},
    {'file': 'get_trading_signal_2399.py', 'name': '2399 映泰'},
    {'file': 'get_trading_signal_2417.py', 'name': '2417 圓剛'},
    {'file': 'get_trading_signal_2431.py', 'name': '2431 聯昌'},
    {'file': 'get_trading_signal_2442.py', 'name': '2442 新美齊'},
    {'file': 'get_trading_signal_2464.py', 'name': '2464 盟立'},
    {'file': 'get_trading_signal_2467.py', 'name': '2467 志聖'},
    {'file': 'get_trading_signal_2472.py', 'name': '2472 立隆電'},
    {'file': 'get_trading_signal_2478.py', 'name': '2478 大毅'},
    {'file': 'get_trading_signal_2484.py', 'name': '2484 希華晶體'},
    {'file': 'get_trading_signal_2489.py', 'name': '2489 瑞軒'},
    {'file': 'get_trading_signal_2492.py', 'name': '2492 華新科'},
    {'file': 'get_trading_signal_2505.py', 'name': '2505 國揚'},
    {'file': 'get_trading_signal_2609.py', 'name': '2609 陽明'},
    {'file': 'get_trading_signal_2610.py', 'name': '2610 中華航空'},
    {'file': 'get_trading_signal_2618.py', 'name': '2618 長榮航空'},
    {'file': 'get_trading_signal_2645.py', 'name': '2645 漢翔'},
    {'file': 'get_trading_signal_2810.py', 'name': '2810 大成鋼'},
    {'file': 'get_trading_signal_2880.py', 'name': '2880 華南金'},
    {'file': 'get_trading_signal_2881.py', 'name': '2881 富邦金'},
    {'file': 'get_trading_signal_2882.py', 'name': '2882 國泰金'},
    {'file': 'get_trading_signal_2886.py', 'name': '2886 兆豐金'},
    {'file': 'get_trading_signal_2892.py', 'name': '2892 第一金'},
    {'file': 'get_trading_signal_2912.py', 'name': '2912 統一超商'},
    {'file': 'get_trading_signal_3016.py', 'name': '3016 嘉晶'},
    {'file': 'get_trading_signal_3036.py', 'name': '3036 文曄'},
    {'file': 'get_trading_signal_3042.py', 'name': '3042 晶技'},
    {'file': 'get_trading_signal_3049.py', 'name': '3049 和鑫'},
    {'file': 'get_trading_signal_3051.py', 'name': '3051 力特'},
    {'file': 'get_trading_signal_3057.py', 'name': '3057 喬鼎'},
    {'file': 'get_trading_signal_3090.py', 'name': '3090 日電貿'},
    {'file': 'get_trading_signal_3092.py', 'name': '3092 鴻碩'},
    {'file': 'get_trading_signal_3149.py', 'name': '3149 正達'},
    {'file': 'get_trading_signal_3167.py', 'name': '3167 渼洋'},
    {'file': 'get_trading_signal_3209.py', 'name': '3209 全科'},
    {'file': 'get_trading_signal_3221.py', 'name': '3221 台嘉碩'},
    {'file': 'get_trading_signal_3234.py', 'name': '3234 光環'},
    {'file': 'get_trading_signal_3293.py', 'name': '3293 鈊象'},
    {'file': 'get_trading_signal_3308.py', 'name': '3308 聯德'},
    {'file': 'get_trading_signal_3338.py', 'name': '3338 泰碩'},
    {'file': 'get_trading_signal_3357.py', 'name': '3357 臺慶科'},
    {'file': 'get_trading_signal_3360.py', 'name': '3360 尚志'},
    {'file': 'get_trading_signal_3374.py', 'name': '3374 精材'},
    {'file': 'get_trading_signal_3402.py', 'name': '3402 聯德控股'},
    {'file': 'get_trading_signal_3432.py', 'name': '3432 台端'},
    {'file': 'get_trading_signal_3485.py', 'name': '3485 敘豐'},
    {'file': 'get_trading_signal_3498.py', 'name': '3498 陽程'},
    {'file': 'get_trading_signal_3532.py', 'name': '3532 台勝科'},
    {'file': 'get_trading_signal_3630.py', 'name': '3630 新巨'},
    {'file': 'get_trading_signal_3645.py', 'name': '3645 達邁'},
    {'file': 'get_trading_signal_3690.py', 'name': '3690 美律'},
    {'file': 'get_trading_signal_3706.py', 'name': '3706 神達'},
    {'file': 'get_trading_signal_3707.py', 'name': '3707 漢磊'},
    {'file': 'get_trading_signal_4142.py', 'name': '4142 國光生技'},
    {'file': 'get_trading_signal_4167.py', 'name': '4167 松瑞藥'},
    {'file': 'get_trading_signal_4533.py', 'name': '4533 協易機'},
    {'file': 'get_trading_signal_4541.py', 'name': '4541 全球傳動'},
    {'file': 'get_trading_signal_4542.py', 'name': '4542 達方'},
    {'file': 'get_trading_signal_4760.py', 'name': '4760 勤凱'},
    {'file': 'get_trading_signal_4900.py', 'name': '4900 富爾特'},
    {'file': 'get_trading_signal_4908.py', 'name': '4908 前鼎'},
    {'file': 'get_trading_signal_4916.py', 'name': '4916 事欣科'},
    {'file': 'get_trading_signal_4919.py', 'name': '4919 新唐科技'},
    {'file': 'get_trading_signal_4927.py', 'name': '4927 泰鼎'},
    {'file': 'get_trading_signal_4966.py', 'name': '4966 譜瑞-KY'},
    {'file': 'get_trading_signal_4973.py', 'name': '4973 廣宇'},
    {'file': 'get_trading_signal_4979.py', 'name': '4979 華星光通'},
    {'file': 'get_trading_signal_5011.py', 'name': '5011 久陽'},
    {'file': 'get_trading_signal_5269.py', 'name': '5269 祥碩'},
    {'file': 'get_trading_signal_5274.py', 'name': '5274 信驊'},
    {'file': 'get_trading_signal_5386.py', 'name': '5386 雷凌科技'},
    {'file': 'get_trading_signal_5475.py', 'name': '5475 德英電子'},
    {'file': 'get_trading_signal_6104.py', 'name': '6104 創源'},
    {'file': 'get_trading_signal_6127.py', 'name': '6127 九豪'},
    {'file': 'get_trading_signal_6133.py', 'name': '6133 金橋'},
    {'file': 'get_trading_signal_6134.py', 'name': '6134 萬旭'},
    {'file': 'get_trading_signal_6135.py', 'name': '6135 新麗'},
    {'file': 'get_trading_signal_6155.py', 'name': '6155 鈦昇'},
    {'file': 'get_trading_signal_6166.py', 'name': '6166 增你強'},
    {'file': 'get_trading_signal_6173.py', 'name': '6173 信昌電'},
    {'file': 'get_trading_signal_6175.py', 'name': '6175 立敦'},
    {'file': 'get_trading_signal_6205.py', 'name': '6205 詮欣'},
    {'file': 'get_trading_signal_6217.py', 'name': '6217 中探針'},
    {'file': 'get_trading_signal_6224.py', 'name': '6224 聚鼎'},
    {'file': 'get_trading_signal_6263.py', 'name': '6263 普萊德'},
    {'file': 'get_trading_signal_6265.py', 'name': '6265 方土昇'},
    {'file': 'get_trading_signal_6282.py', 'name': '6282 康舒'},
    {'file': 'get_trading_signal_6284.py', 'name': '6284 佳邦'},
    {'file': 'get_trading_signal_6344.py', 'name': '6344 萬年清'},
    {'file': 'get_trading_signal_6415.py', 'name': '6415 矽力-KY'},
    {'file': 'get_trading_signal_6423.py', 'name': '6423 精測'},
    {'file': 'get_trading_signal_6426.py', 'name': '6426 統新'},
    {'file': 'get_trading_signal_6456.py', 'name': '6456 GreenPower'},
    {'file': 'get_trading_signal_6457.py', 'name': '6457 醣聯'},
    {'file': 'get_trading_signal_6485.py', 'name': '6485 點序'},
    {'file': 'get_trading_signal_6530.py', 'name': '6530 創威'},
    {'file': 'get_trading_signal_6570.py', 'name': '6570 維田'},
    {'file': 'get_trading_signal_6574.py', 'name': '6574 霖揚'},
    {'file': 'get_trading_signal_6584.py', 'name': '6584 申豐'},
    {'file': 'get_trading_signal_6592.py', 'name': '6592 和潤企業'},
    {'file': 'get_trading_signal_6603.py', 'name': '6603 富奇想'},
    {'file': 'get_trading_signal_6658.py', 'name': '6658 聯策'},
    {'file': 'get_trading_signal_6691.py', 'name': '6691 長天科技'},
    {'file': 'get_trading_signal_6727.py', 'name': '6727 亞泰金屬'},
    {'file': 'get_trading_signal_6829.py', 'name': '6829 勝麗'},
    {'file': 'get_trading_signal_6831.py', 'name': '6831 騰雲'},
    {'file': 'get_trading_signal_6834.py', 'name': '6834 新纖維'},
    {'file': 'get_trading_signal_6835.py', 'name': '6835 複合互連'},
    {'file': 'get_trading_signal_6862.py', 'name': '6862 勝麗-KY'},
    {'file': 'get_trading_signal_6903.py', 'name': '6903 信強'},
    {'file': 'get_trading_signal_6944.py', 'name': '6944 譜力'},
    {'file': 'get_trading_signal_6949.py', 'name': '6949 連展投控'},
    {'file': 'get_trading_signal_6958.py', 'name': '6958 無限金融'},
    {'file': 'get_trading_signal_6988.py', 'name': '6988 正基'},
    {'file': 'get_trading_signal_6994.py', 'name': '6994 環科'},
    {'file': 'get_trading_signal_7728.py', 'name': '7728 中磊'},
    {'file': 'get_trading_signal_7744.py', 'name': '7744 致伸'},
    {'file': 'get_trading_signal_7795.py', 'name': '7795 億泰'},
    {'file': 'get_trading_signal_7805.py', 'name': '7805 台灣外包'},
    {'file': 'get_trading_signal_8033.py', 'name': '8033 雷虎'},
    {'file': 'get_trading_signal_8038.py', 'name': '8038 長園科'},
    {'file': 'get_trading_signal_8039.py', 'name': '8039 台虹'},
    {'file': 'get_trading_signal_8042.py', 'name': '8042 金山電'},
    {'file': 'get_trading_signal_8043.py', 'name': '8043 蜜望寶'},
    {'file': 'get_trading_signal_8074.py', 'name': '8074 特技'},
    {'file': 'get_trading_signal_8086.py', 'name': '8086 宏捷科'},
    {'file': 'get_trading_signal_8271.py', 'name': '8271 宇瞻'},
    {'file': 'get_trading_signal_8292.py', 'name': '8292 邦特'},
    {'file': 'get_trading_signal_8341.py', 'name': '8341 日友'},
    {'file': 'get_trading_signal_8358.py', 'name': '8358 金居'},
    {'file': 'get_trading_signal_8377.py', 'name': '8377 建漢'},
    {'file': 'get_trading_signal_8431.py', 'name': '8431 匯僑'},
    {'file': 'get_trading_signal_8450.py', 'name': '8450 霸機科'},
    {'file': 'get_trading_signal_8462.py', 'name': '8462 柯訊'},
    {'file': 'get_trading_signal_9136.py', 'name': '9136 巨騰'},
    {'file': 'get_trading_signal_9888.py', 'name': '9888 碧桂園服務'},
    {'file': 'get_trading_signal_9933.py', 'name': '9933 中鼎'},
    {'file': 'get_trading_signal_9984.py', 'name': '9984 軟銀'},
]

def get_ticker_from_code(stock_code):
    """从股票代码获取完整的ticker符号"""
    # Tokyo market stocks
    T_STOCKS = {'3449'}

    # TWO exchange stocks
    TWO_STOCKS = {'3498', '3615', '4533', '4577', '4768', '4908', '4991', '5011',
                  '6134', '6187', '6220', '6530', '6877', '7805', '8086', '8908', '8917', '8927',
                  '6274', '1785', '4749', '3131', '6683', '3363', '3081', '6510', '1815',  # added TWO-market stocks
                  '8069', '6223', '5483', '6163', '7709', '7717',
                  '3260', '3491', '5371', '3105', '4971',
                  '8064', '3163', '3455',
                  '3680', '4772', '6788', '7703', '8147', '8071',
                  '8027', '5351', '7734', '7751', '6138',
                  '1569', '1595', '4951',
                  '6234', '6488', '6207',
                  '6640', '6664',
                  # previously failed as .TW — correct excha
                  # 1nge is TWO
                  '1815', '3147', '3152', '3236', '3264', '3265', '3577', '3580', '3581', '3587',
                  '3609', '3624', '3663', '3691', '5289', '6146', '6147', '6204', '6432', '6538',
                  '6643', '6980', '8044', '8291', '8299', '8455',  # verified TWO stocks
                  # ── Batch 9 additions (2026-06-09) ──
                  '3221', '3234', '3357', '3402', '3485', '3498', '4542', '4760',
                  '4966', '4979', '5386', '5475', '6127', '6166', '6173', '6175',
                  '6217', '6263', '6284', '6344', '6426', '6570', '6574', '6584',
                  '6691', '6727', '6903', '7744', '8042', '8043','7610'}

    if stock_code in T_STOCKS:
        return f"{stock_code}.T"
    elif stock_code in TWO_STOCKS:
        return f"{stock_code}.TWO"
    else:
        return f"{stock_code}.TW"

def fetch_stock_metrics(stock_code):
    """Fetch all metrics matching the western stocks template format"""
    _none17 = (None,) * 21
    try:
        ticker = get_ticker_from_code(stock_code)
        beta = None
        peg = None
        rev_growth = None
        eps_growth = None
        inst_owned = None
        ma50_info = None
        ma200_info = None
        raw_rev = None
        raw_eps = None
        raw_inst = None

        try:
            info = yf.Ticker(ticker).info
            beta = info.get('beta')
            if beta is not None:
                beta = round(float(beta), 2)
            raw_peg = info.get('pegRatio')
            if raw_peg is not None:
                peg = round(float(raw_peg), 2)
            raw_rev = info.get('revenueGrowth')
            if raw_rev is not None:
                rev_growth = round(float(raw_rev) * 100, 1)
            raw_eps = info.get('earningsGrowth')
            if raw_eps is not None:
                eps_growth = round(float(raw_eps) * 100, 1)
            raw_inst = info.get('heldPercentInstitutions') or info.get('institutionPercentHeld')
            if raw_inst is not None:
                inst_owned = round(float(raw_inst) * 100, 1)
            v = info.get('fiftyDayAverage')
            if v is not None:
                ma50_info = round(float(v), 2)
            v = info.get('twoHundredDayAverage')
            if v is not None:
                ma200_info = round(float(v), 2)
        except:
            pass

        df = yf.download(ticker, period='3mo', progress=False, auto_adjust=True)
        if df.empty or len(df) < 10:
            return _none17

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)

        # Compute volume signal early (needs Open column before dropna on close)
        volume_signal = None
        if len(df) >= 2 and 'Open' in df.columns:
            try:
                lc = float(df['Close'].iloc[-1])
                lo = float(df['Open'].iloc[-1])
                lv = float(df['Volume'].iloc[-1])
                pv = float(df['Volume'].iloc[-2])
                if lv > pv and lc > lo:
                    volume_signal = "✅ 放量收阳"
            except Exception:
                pass

        close = df['Close'].dropna()
        if len(close) < 10:
            return _none17

        # Daily change %
        change_pct = None
        if len(close) >= 2:
            val = float((close.iloc[-1] - close.iloc[-2]) / close.iloc[-2] * 100)
            if not np.isnan(val):
                change_pct = round(val, 2)

        # Current price and volume
        current_price = float(close.iloc[-1]) if not np.isnan(float(close.iloc[-1])) else None
        vol_series = df['Volume'].dropna()
        current_volume = int(vol_series.iloc[-1]) if len(vol_series) > 0 else None
        trading_value = round(current_price * current_volume, 2) if current_price and current_volume else None
        # Sector Momentum: 20-day price change %
        sector_momentum = None
        if len(close) >= 20:
            val = float((close.iloc[-1] - close.iloc[-20]) / close.iloc[-20] * 100)
            if not np.isnan(val):
                sector_momentum = round(val, 2)

        # Short-term Volatility
        short_vol = None
        returns = close.pct_change().dropna()
        if len(returns) >= 10:
            val = float(returns.tail(20).std() * (252 ** 0.5) * 100)
            if not np.isnan(val):
                short_vol = round(val, 1)

        # Breakout Probability
        breakout_prob = None
        if len(close) >= 20:
            bb_mid = close.rolling(20).mean()
            bb_std_val = close.rolling(20).std()
            bb_upper = bb_mid + 2 * bb_std_val
            bb_lower = bb_mid - 2 * bb_std_val
            bb_range = float(bb_upper.iloc[-1] - bb_lower.iloc[-1])
            if bb_range > 0 and not np.isnan(bb_range):
                bb_pos = float((close.iloc[-1] - bb_lower.iloc[-1]) / bb_range * 100)
                if not np.isnan(bb_pos):
                    vol_mean = float(vol_series.tail(20).mean()) if len(vol_series) >= 20 else float(vol_series.mean())
                    vol_ratio = float(vol_series.iloc[-1] / vol_mean) if vol_mean > 0 else 1.0
                    if not np.isnan(vol_ratio):
                        breakout_prob = round(min(100, max(0, bb_pos * 0.5 + min(vol_ratio, 3) / 3 * 50)), 1)

        # RSI (14)
        rsi_val = None
        if len(close) >= 15:
            delta = close.diff()
            gain = delta.where(delta > 0, 0).rolling(14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
            rs = gain / (loss + 1e-10)
            rsi_series = 100 - (100 / (1 + rs))
            v = float(rsi_series.iloc[-1])
            if not np.isnan(v):
                rsi_val = round(v, 1)

        # MA Trend (50MA vs 200MA)
        ma_trend = None
        if ma50_info is not None and ma200_info is not None:
            ma_trend = 1 if ma50_info > ma200_info else 0

        # Composite score
        score = 0.0
        if peg is not None and peg < 1:
            score += 2
        if raw_rev is not None:
            score += float(raw_rev) * 5
        if raw_eps is not None:
            score += float(raw_eps) * 3
        if raw_inst is not None:
            score += float(raw_inst)
        if ma_trend:
            score += 1
        score = round(score, 2)

        # ── MACD 柱狀體收腳 detection (reuse existing df) ────────────────────
        macd_foot = False
        macd_gap_up = False
        macd_shrink = 0.0
        try:
            if len(close) >= 40:
                ema_f = close.ewm(span=12, adjust=False).mean()
                ema_s = close.ewm(span=26, adjust=False).mean()
                ml    = ema_f - ema_s
                ms    = ml.ewm(span=9, adjust=False).mean()
                mh    = ml - ms          # histogram
                if len(mh) >= 3:
                    h0 = float(mh.iloc[-1])
                    h1 = float(mh.iloc[-2])
                    h2 = float(mh.iloc[-3])
                    if h0 < 0 and h1 < 0 and h2 < 0 and abs(h0) < abs(h1):
                        shrink = (abs(h1) - abs(h0)) / (abs(h1) + 1e-10) * 100
                        if shrink >= 10:
                            macd_foot   = True
                            macd_shrink = round(shrink, 1)
                            # Gap up: today's open > yesterday's high
                            if 'Open' in df.columns and 'High' in df.columns:
                                t_open  = float(df['Open'].dropna().iloc[-1])
                                y_high  = float(df['High'].dropna().iloc[-2])
                                macd_gap_up = t_open > y_high
        except Exception:
            pass

        return (change_pct, current_price, current_volume, trading_value,
                sector_momentum, beta, short_vol, breakout_prob,
                peg, rev_growth, eps_growth, inst_owned,
                ma50_info, ma200_info, rsi_val, ma_trend, score, volume_signal,
                macd_foot, macd_gap_up, macd_shrink)

    except Exception:
        return _none17

def analyze_volume_and_obv(stock_code):
    """分析资金流入和OBV趋势"""
    try:
        ticker = get_ticker_from_code(stock_code)
        df = yf.download(ticker, period='3mo', progress=False)

        if df.empty or len(df) < 5:
            return None, None

        # Flatten multi-index columns if necessary
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)

        # 计算OBV
        df['price_change'] = df['Close'].diff()
        df['obv'] = (np.sign(df['price_change']) * df['Volume']).fillna(0).cumsum()

        # 获取最近的数据
        latest = df.iloc[-1]
        prev = df.iloc[-2] if len(df) > 1 else None

        # 检查是否放量收阳
        volume_expansion = False
        positive_close = False

        if prev is not None:
            # 放量：今日成交量 > 前一日成交量
            volume_expansion = latest['Volume'] > prev['Volume']
            # 收阳：收盘价 > 开盘价
            positive_close = latest['Close'] > latest['Open']

        volume_signal = "✅ 放量收阳" if (volume_expansion and positive_close) else ""

        # 检查OBV趋势 (最近5天)
        obv_bullish = False
        if len(df) >= 5:
            obv_recent = df['obv'].iloc[-5:].values
            # 简单判断：OBV整体上升
            obv_bullish = obv_recent[-1] > obv_recent[0]

        obv_signal = "✅ bullish" if obv_bullish else ""

        return volume_signal, obv_signal

    except Exception as e:
        return None, None

def parse_accuracy_from_output(output):
    """从输出中解析模型准确率"""
    import re
    patterns = [
        # TW scripts: "AI準確度: 72.7/100"
        r'AI準確度[：:]\s*(\d+\.?\d*)\s*/\s*100',
        # Western/generic percentage patterns
        r'(?:accuracy|准确率|準確率)[^\d]*(\d+\.?\d*)\s*%',
        r'(?:accuracy|准确率|準確率)[^\d]*(\d+\.\d+)',
        r'(\d+\.?\d*)\s*%\s*(?:accuracy|准确率|準確率)',
    ]
    for pat in patterns:
        m = re.search(pat, output, re.IGNORECASE)
        if m:
            try:
                val = float(m.group(1))
                if val > 1:
                    return round(val, 2)
                else:
                    return round(val * 100, 2)
            except Exception:
                pass
    return None

def parse_signal_from_output(output):
    """从输出中解析交易信号"""
    output_lower = output.lower()
    if ("🟢 買入信號" in output or "买入信号" in output_lower
            or "(BUY)" in output or "买入 (buy)" in output_lower):
        return "BUY"
    elif ("🔴 賣出信號" in output or "卖出信号" in output_lower
            or "(SELL)" in output or "卖出 (sell)" in output_lower):
        return "SELL"
    elif "(HOLD" in output or "持有 (hold" in output_lower or "强势持有" in output:
        return "HOLD"
    elif "(WAIT)" in output or "观望 (wait)" in output_lower:
        return "WAIT"
    elif "(BUYSELL)" in output or "buysell" in output_lower:
        return "BUYSELL"
    return "UNKNOWN"

def run_signal_and_capture(script_file, stock_name):
    """运行单个交易信号生成器并捕获输出"""
    print(f"\n进度: 运行 {stock_name}...", flush=True)

    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        script_path = os.path.join(script_dir, script_file)

        # Use the same Python interpreter that's running this script
        python_executable = sys.executable
        
        result = subprocess.run(
            [python_executable, script_path],
            capture_output=True,
            text=True,
            timeout=180,
            encoding='utf-8',
            errors='ignore',
            cwd=script_dir
        )

        if result.returncode == 0:
            stdout = result.stdout
            _err = ["无法获取数据", "模型加载失败", "信号生成失败", "下载失败", "No data found"]
            _sig = ["(BUY)", "(SELL)", "(HOLD", "(WAIT)", "(BUYSELL)", "買入信號", "卖出信号", "买入 (buy)", "卖出 (sell)"]
            if any(e in stdout for e in _err) and not any(s in stdout for s in _sig):
                return False, f"[ERROR] Script ran but no signal produced\n{stdout[-500:]}"
            return True, stdout
        else:
            return False, f"[ERROR]\n{result.stderr}"

    except subprocess.TimeoutExpired:
        return False, "[ERROR] 超时 (180秒)"
    except Exception as e:
        return False, f"[ERROR] {str(e)}"

def write_output_to_sheet(ws, stock_name, output_text, success):
    """将输出写入工作表"""
    # 设置标题
    ws['A1'] = stock_name
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4" if success else "C55A11",
                                 end_color="4472C4" if success else "C55A11",
                                 fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')

    # 设置状态
    status = "✅ 成功" if success else "❌ 失败"
    ws['A2'] = "状态"
    ws['B2'] = status
    ws['A2'].font = Font(bold=True)

    # 写入输出内容
    ws['A3'] = "交易信号输出"
    ws['A3'].font = Font(bold=True)
    ws.merge_cells('A3:B3')

    # 将输出按行分割并写入
    import re as _re
    _illegal = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
    lines = output_text.split('\n')
    for idx, line in enumerate(lines, start=4):
        if idx > 1000:  # 限制最多1000行以防止Excel过大
            ws[f'A{idx}'] = "... (输出过长，已截断)"
            break
        ws[f'A{idx}'] = _illegal.sub('', line)
        ws.merge_cells(f'A{idx}:B{idx}')

    # 调整列宽
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 50

def create_summary_sheet(wb, results):
    """创建汇总工作表 (matching western stocks template format)"""
    ws = wb.create_sheet("汇总", 0)

    # 标题
    ws['A1'] = "台股交易信号汇总"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:AB1')

    # 统计信息
    ws['A3'] = "生成时间:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A4'] = "总股票数:"
    ws['B4'] = len(results)
    ws['A5'] = "成功数量:"
    ws['B5'] = sum(1 for r in results if r['success'])
    ws['A6'] = "失败数量:"
    ws['B6'] = sum(1 for r in results if not r['success'])

    for cell in ['A3', 'A4', 'A5', 'A6']:
        ws[cell].font = Font(bold=True)

    # 列标题 — matches western template
    headers = {
        'A8': '序号',
        'B8': '股票代码',
        'C8': '股票名称',
        'D8': 'change',
        'E8': 'AI信号',
        'F8': 'volumn',
        'G8': 'price',
        'H8': '交易金額',
        'I8': 'Sector Momentum (20d%)',
        'J8': 'Beta',
        'K8': 'Short-term Volatility',
        'L8': 'Breakout Probability',
        'M8': 'PEG Ratio',
        'N8': 'Revenue Growth',
        'O8': 'EPS Growth',
        'P8': 'Institution Ownership',
        'Q8': '50 Day MA',
        'R8': '200 Day MA',
        'S8': 'RSI (14)',
        'T8': 'MA Trend',
        'U8': 'Score',
        'V8': '即时资金流入',
        'W8': 'AI准确率',
        'X8': 'MACD收腳',
        'Y8': '跳空缺口',
        'Z8':  'MACD格局',
        'AA8': 'Hybrid PPO',
        'AB8': 'Hybrid Action',
    }
    for cell, label in headers.items():
        ws[cell] = label
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # 写入股票列表
    for idx, result in enumerate(results, start=9):
        ws[f'A{idx}'] = idx - 8
        parts = result['name'].split(maxsplit=1)
        stock_code = parts[0]
        stock_name = parts[1] if len(parts) > 1 else ""

        ws[f'B{idx}'] = stock_code
        ws[f'C{idx}'] = stock_name
        sheet_name = result.get('sheet_name', '')
        if sheet_name:
            ws[f'C{idx}'].hyperlink = f"#'{sheet_name}'!A1"
            ws[f'C{idx}'].font = Font(color="0563C1", underline="single")

        # AI 信号
        signal = parse_signal_from_output(result['output']) if result['success'] else "N/A"
        ws[f'E{idx}'] = signal
        _signal_colors = {
            "BUY":     ("C6EFCE", "006100"),
            "SELL":    ("FFC7CE", "9C0006"),
            "HOLD":    ("DDEBF7", "2E75B6"),
            "WAIT":    ("FFEB9C", "9C6500"),
            "BUYSELL": ("FFD966", "7F4C00"),
            "UNKNOWN": ("F2F2F2", "808080"),
        }
        if signal in _signal_colors:
            bg, fg = _signal_colors[signal]
            ws[f'E{idx}'].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws[f'E{idx}'].font = Font(color=fg, bold=(signal not in ("UNKNOWN",)))

        if result['success']:
            print(f"   📊 获取 {stock_code} 指标数据...", flush=True)
            time.sleep(1)
            metrics = fetch_stock_metrics(stock_code)
            (change_pct, current_price, current_volume, trading_value,
             sector_momentum, beta, short_vol, breakout_prob,
             peg, rev_growth, eps_growth, inst_owned,
             ma50, ma200, rsi_val, ma_trend, score, volume_signal,
             macd_foot, macd_gap_up, macd_shrink) = metrics

            # D: change
            if change_pct is not None:
                ws[f'D{idx}'] = f"{change_pct:+.2f}%"
                if change_pct > 0:
                    ws[f'D{idx}'].font = Font(color="006100")
                elif change_pct < 0:
                    ws[f'D{idx}'].font = Font(color="9C0006")

            # F: volume
            if current_volume is not None:
                ws[f'F{idx}'] = current_volume
                ws[f'F{idx}'].number_format = '#,##0'

            # G: price
            if current_price is not None:
                ws[f'G{idx}'] = current_price
                ws[f'G{idx}'].number_format = 'NT$#,##0.00'

            # H: 交易金額
            if trading_value is not None:
                ws[f'H{idx}'] = trading_value
                ws[f'H{idx}'].number_format = 'NT$#,##0.00'

            # I: Sector Momentum
            if sector_momentum is not None:
                ws[f'I{idx}'] = f"{sector_momentum:+.2f}%"
                if sector_momentum > 5:
                    ws[f'I{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'I{idx}'].font = Font(color="006100")
                elif sector_momentum < -5:
                    ws[f'I{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'I{idx}'].font = Font(color="9C0006")

            # J: Beta
            if beta is not None:
                ws[f'J{idx}'] = beta
                if beta > 1.5:
                    ws[f'J{idx}'].font = Font(color="9C0006")
                elif beta < 0.5:
                    ws[f'J{idx}'].font = Font(color="006100")

            # K: Short-term Volatility
            if short_vol is not None:
                ws[f'K{idx}'] = f"{short_vol:.1f}%"
                if short_vol > 50:
                    ws[f'K{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'K{idx}'].font = Font(color="9C0006")

            # L: Breakout Probability
            if breakout_prob is not None:
                ws[f'L{idx}'] = f"{breakout_prob:.1f}%"
                if breakout_prob >= 70:
                    ws[f'L{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'L{idx}'].font = Font(color="006100", bold=True)
                elif breakout_prob >= 50:
                    ws[f'L{idx}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    ws[f'L{idx}'].font = Font(color="9C6500")

            # M: PEG Ratio
            if peg is not None:
                ws[f'M{idx}'] = peg
                if peg < 1:
                    ws[f'M{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'M{idx}'].font = Font(color="006100")
                elif peg > 2:
                    ws[f'M{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'M{idx}'].font = Font(color="9C0006")

            # N: Revenue Growth
            if rev_growth is not None:
                ws[f'N{idx}'] = f"{rev_growth:+.1f}%"
                if rev_growth > 10:
                    ws[f'N{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'N{idx}'].font = Font(color="006100")
                elif rev_growth < 0:
                    ws[f'N{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'N{idx}'].font = Font(color="9C0006")

            # O: EPS Growth
            if eps_growth is not None:
                ws[f'O{idx}'] = f"{eps_growth:+.1f}%"
                if eps_growth > 10:
                    ws[f'O{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'O{idx}'].font = Font(color="006100")
                elif eps_growth < 0:
                    ws[f'O{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'O{idx}'].font = Font(color="9C0006")

            # P: Institution Ownership
            if inst_owned is not None:
                ws[f'P{idx}'] = f"{inst_owned:.1f}%"

            # Q: 50 Day MA
            if ma50 is not None:
                ws[f'Q{idx}'] = ma50

            # R: 200 Day MA
            if ma200 is not None:
                ws[f'R{idx}'] = ma200

            # S: RSI (14)
            if rsi_val is not None:
                ws[f'S{idx}'] = rsi_val
                if rsi_val >= 70:
                    ws[f'S{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'S{idx}'].font = Font(color="9C0006", bold=True)
                elif rsi_val <= 30:
                    ws[f'S{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'S{idx}'].font = Font(color="006100", bold=True)

            # T: MA Trend
            if ma_trend is not None:
                ws[f'T{idx}'] = "Bullish" if ma_trend == 1 else "Bearish"
                if ma_trend == 1:
                    ws[f'T{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'T{idx}'].font = Font(color="006100")
                else:
                    ws[f'T{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'T{idx}'].font = Font(color="9C0006")

            # U: Score
            if score is not None:
                ws[f'U{idx}'] = score
                if score >= 3:
                    ws[f'U{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'U{idx}'].font = Font(color="006100", bold=True)
                elif score < 0:
                    ws[f'U{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'U{idx}'].font = Font(color="9C0006", bold=True)

            # V: 即时资金流入 (volume signal + RSI)
            parts = []
            if volume_signal:
                parts.append(volume_signal)
            if rsi_val is not None:
                parts.append(f"RSI:{rsi_val}")
            if parts:
                ws[f'V{idx}'] = " | ".join(parts)
                if volume_signal:
                    ws[f'V{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'V{idx}'].font = Font(color="006100", bold=True)

            # W: AI准确率
            accuracy = parse_accuracy_from_output(result['output'])
            if accuracy is not None:
                ws[f'W{idx}'] = f"{accuracy:.2f}%"
                if accuracy >= 65:
                    ws[f'W{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'W{idx}'].font = Font(color="006100", bold=True)
                elif accuracy < 50:
                    ws[f'W{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'W{idx}'].font = Font(color="9C0006")

            # X: MACD收腳
            if macd_foot:
                gap_txt = " ✅跳空" if macd_gap_up else ""
                ws[f'X{idx}'] = f"✅ {macd_shrink:.0f}%{gap_txt}"
                ws[f'X{idx}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                ws[f'X{idx}'].font = Font(color="7F4C00", bold=True)
            else:
                ws[f'X{idx}'] = "—"

            # Y: 跳空缺口
            if macd_gap_up:
                ws[f'Y{idx}'] = "✅ 跳空"
                ws[f'Y{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'Y{idx}'].font = Font(color="006100", bold=True)
            else:
                ws[f'Y{idx}'] = "—"

            # Z: MACD格局 (inline daily-only proxy using macd from fetch_stock_metrics)
            # Use rsi_val as proxy availability check; compute inline from metrics we already have
            try:
                _macd_line = metrics[4] if False else None  # placeholder
                # Derive from existing data already in summary context
                # We compute bull status from the macd_foot + volume_signal context:
                if macd_foot and macd_shrink and macd_shrink > 0:
                    bull_status = "✅ 強勢整理"
                    ws[f'Z{idx}'] = bull_status
                    ws[f'Z{idx}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    ws[f'Z{idx}'].font = Font(color="2E75B6", bold=True)
                elif not macd_foot:
                    # hist not negative or not shrinking → could be 完美多頭 or 日線轉負
                    # Use volume_signal as indirect indicator of strength
                    ws[f'Z{idx}'] = "🟢 多頭" if volume_signal else "—"
                    if volume_signal:
                        ws[f'Z{idx}'].font = Font(color="006100")
            except Exception:
                ws[f'Z{idx}'] = "—"

            # AA/AB: Hybrid PPO signal
            try:
                from hybrid_signal_helper import get_hybrid_signal, find_hybrid_model
                if find_hybrid_model(stock_code):
                    _df = yf.download(stock_code if '.' in stock_code else
                                      get_ticker_from_code(stock_code), period='300d',
                                      progress=False, auto_adjust=True)
                    if not _df.empty:
                        if isinstance(_df.columns, pd.MultiIndex):
                            _df.columns = _df.columns.droplevel(1)
                        _df = _df.rename(columns={'Close':'close','Volume':'volume',
                                                   'Open':'open','High':'high','Low':'low'})
                        _h = get_hybrid_signal(stock_code, _df)
                        _sig = _h.get('signal', '—')
                        _act = _h.get('action', 0.0)
                        ws[f'AA{idx}'] = _sig
                        ws[f'AB{idx}'] = f"{_act:+.3f}"
                        _colors = {'BUY':('C6EFCE','006100'), 'SELL':('FFC7CE','9C0006'),
                                   'HOLD':('DDEBF7','2E75B6')}
                        if _sig in _colors:
                            bg, fg = _colors[_sig]
                            ws[f'AA{idx}'].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                            ws[f'AA{idx}'].font = Font(color=fg, bold=True)
                else:
                    ws[f'AA{idx}'] = "—"; ws[f'AB{idx}'] = "—"
            except Exception:
                ws[f'AA{idx}'] = "—"; ws[f'AB{idx}'] = "—"

        else:
            for col in ['D','F','G','H','I','J','K','L','M','N','O','P','Q','R',
                        'S','T','U','V','W','X','Y','Z','AA','AB']:
                ws[f'{col}{idx}'] = ""

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 22
    ws.column_dimensions['Q'].width = 14
    ws.column_dimensions['R'].width = 14
    ws.column_dimensions['S'].width = 12
    ws.column_dimensions['T'].width = 12
    ws.column_dimensions['U'].width = 10
    ws.column_dimensions['V'].width = 24
    ws.column_dimensions['W'].width = 14
    ws.column_dimensions['X'].width = 16
    ws.column_dimensions['Y'].width = 12
    ws.column_dimensions['Z'].width  = 14
    ws.column_dimensions['AA'].width = 12
    ws.column_dimensions['AB'].width = 12

if __name__ == "__main__":
    import sys as _sys
    txt_only = '--txt-only' in _sys.argv

    print("=" * 100)
    print("批量运行所有台股交易信号生成器" + (" (仅文字输出)" if txt_only else " (输出到Excel)"))
    print("=" * 100)
    print(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"总共 {len(SIGNAL_SCRIPTS)} 个台股")
    print("=" * 100)

    if not txt_only:
        wb = Workbook()
        wb.remove(wb.active)

    results = []
    success_count = 0
    failed_stocks = []
    interrupted = False

    try:
        for i, script in enumerate(SIGNAL_SCRIPTS, 1):
            print(f"\n[{i}/{len(SIGNAL_SCRIPTS)}] 运行 {script['name']}...", flush=True)

            success, output = run_signal_and_capture(script['file'], script['name'])

            if not txt_only:
                sheet_name = script['name'].replace('/', '_').replace('\\', '_').replace('*', '_')
                sheet_name = sheet_name.replace('?', '_').replace(':', '_').replace('[', '_').replace(']', '_')
                if len(sheet_name) > 31:
                    sheet_name = script['name'].split()[0]
                actual_sheet_name = sheet_name
                try:
                    ws = wb.create_sheet(sheet_name)
                    write_output_to_sheet(ws, script['name'], output, success)
                except Exception as e:
                    print(f"   ⚠️  创建工作表失败: {e}")
                    actual_sheet_name = f"Stock_{i}"
                    ws = wb.create_sheet(actual_sheet_name)
                    write_output_to_sheet(ws, script['name'], output, success)

            results.append({'name': script['name'], 'success': success, 'output': output,
                            'sheet_name': actual_sheet_name if not txt_only else ''})

            if success:
                success_count += 1
                print(f"   ✅ 成功")
            else:
                failed_stocks.append(script['name'])
                print(f"   ❌ 失败")

    except KeyboardInterrupt:
        print("\n\n⚠️  检测到中断 (Ctrl+C)！正在保存已完成的结果...")
        interrupted = True

    print("\n" + "=" * 100)
    print("批量运行被中断!" if interrupted else "批量运行完成!")
    print("=" * 100)
    print(f"成功运行: {success_count}/{len(results)} (总共 {len(SIGNAL_SCRIPTS)} 个)")
    print(f"失败数量: {len(failed_stocks)}")

    if failed_stocks:
        print(f"\n失败的股票:")
        for stock in failed_stocks:
            print(f"   - {stock}")

    if not txt_only and results:
        create_summary_sheet(wb, results)
        timestamp = datetime.now().strftime('%Y%m%d%H%M')
        output_filename = f'taiwan_signals_output_{timestamp}{"_PARTIAL" if interrupted else ""}.xlsx'
        output_path = os.path.join(SCRIPT_DIR, output_filename)
        wb.save(output_path)
        print(f"\n✅ Excel文件已保存: {output_path}")

    if results:
        print("\n所有台股信号生成完成!")
    else:
        print("\n❌ 没有生成任何结果")

