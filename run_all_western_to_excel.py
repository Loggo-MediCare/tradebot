"""
批量运行所有西方股票的交易信号生成器 (输出到Excel)
================================
自动运行所有已训练的美股和欧股模型的交易信号
每个股票的输出保存到单独的工作表
"""

import asyncio
import sys
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import random
import yfinance as yf
import numpy as np
import pandas as pd

# 修复 Windows 编码问题
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 获取脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Max signal scripts running in parallel (keeps yfinance rate-limits happy)
CONCURRENT_LIMIT = 3

FX_TO_USD_CACHE = {'USD': 1.0}


def _safe_float(value):
    try:
        if value is None:
            return None
        value = float(value)
        if np.isnan(value):
            return None
        return value
    except Exception:
        return None


def _as_series(df, column):
    values = df[column]
    if isinstance(values, pd.DataFrame):
        values = values.iloc[:, 0]
    return values


def infer_currency_from_ticker(ticker_symbol):
    """Best-effort currency fallback when Yahoo info omits currency."""
    ticker = str(ticker_symbol).upper()
    if ticker.endswith('.DE'):
        return 'EUR'
    if ticker.endswith('.HK'):
        return 'HKD'
    if ticker.endswith('.T'):
        return 'JPY'
    if ticker.endswith('.KS'):
        return 'KRW'
    if ticker.endswith('.TW') or ticker.endswith('.TWO'):
        return 'TWD'
    return 'USD'


def normalize_ticker_for_yfinance(stock_code):
    """Match the ticker format expected by Yahoo Finance."""
    ticker = str(stock_code).strip()
    if ticker.endswith('.HK') and ticker[0] == '0':
        stripped = ticker.lstrip('0')
        return stripped if stripped else ticker
    return ticker


def split_stock_name(display_name):
    parts = str(display_name).split(maxsplit=1)
    stock_code = parts[0] if parts else ''
    stock_name = parts[1] if len(parts) > 1 else ''
    return stock_code, stock_name


def get_fx_to_usd(currency):
    """Return USD per 1 unit of currency using Yahoo Finance FX pairs."""
    if not currency:
        return 1.0

    ccy = str(currency).strip().upper()
    if ccy in FX_TO_USD_CACHE:
        return FX_TO_USD_CACHE[ccy]

    # London stocks sometimes quote in pence. Convert pence -> GBP -> USD.
    if ccy in ('GBX', 'GBPENCE', 'GBP PENCE'):
        gbp_rate = get_fx_to_usd('GBP')
        rate = gbp_rate / 100 if gbp_rate else None
        FX_TO_USD_CACHE[ccy] = rate
        return rate

    candidates = [
        (f'{ccy}USD=X', False),  # e.g. EURUSD=X = USD per EUR
        (f'USD{ccy}=X', True),   # e.g. USDJPY=X = JPY per USD, invert it
    ]

    for fx_ticker, invert in candidates:
        try:
            fx_df = yf.download(fx_ticker, period='5d', progress=False, auto_adjust=True)
            if fx_df.empty or 'Close' not in fx_df.columns:
                continue
            if isinstance(fx_df.columns, pd.MultiIndex):
                fx_df.columns = fx_df.columns.droplevel(1)
            close = _as_series(fx_df, 'Close').dropna()
            if close.empty:
                continue
            raw_rate = _safe_float(close.iloc[-1])
            if raw_rate and raw_rate > 0:
                rate = (1.0 / raw_rate) if invert else raw_rate
                FX_TO_USD_CACHE[ccy] = rate
                return rate
        except Exception:
            continue

    FX_TO_USD_CACHE[ccy] = None
    return None

# 西方股票信号生成器 (US + EU)
SIGNAL_SCRIPTS = [
    # Top Priority US Stocks (User Preferred Order)
    {'file': 'get_trading_signal_nvda.py', 'name': 'NVDA NVIDIA'},
    {'file': 'get_trading_signal_tsla.py', 'name': 'TSLA Tesla'},
    {'file': 'get_trading_signal_googl.py', 'name': 'GOOGL Alphabet'},
    {'file': 'get_trading_signal_avgo.py', 'name': 'AVGO Broadcom'},
    {'file': 'get_trading_signal_mu.py', 'name': 'MU Micron'},
    
    {'file': 'get_trading_signal_sndk.py', 'name': 'SNDK SanDisk'},
    {'file': 'get_trading_signal_msft.py', 'name': 'MSFT Microsoft'},
    {'file': 'get_trading_signal_amzn.py', 'name': 'AMZN Amazon'},
    {'file': 'get_trading_signal_aapl.py', 'name': 'AAPL Apple'},
    {'file': 'get_trading_signal_meta.py', 'name': 'META Meta Platforms'},
    {'file': 'get_trading_signal_pltr.py', 'name': 'PLTR Palantir'},
    {'file': 'get_trading_signal_tsm.py', 'name': 'TSM TSMC ADR'},
    {'file': 'get_trading_signal_wdc.py', 'name': 'WDC Western Digital'},

    # Other US Stocks
    {'file': 'get_trading_signal_aeva.py', 'name': 'AEVA Aeva Technologies Inc'},
    {'file': 'get_trading_signal_alab.py', 'name': 'ALAB Astera Labs Inc'},
    {'file': 'get_trading_signal_amkr.py', 'name': 'AMKR Amkor'},
    {'file': 'get_trading_signal_avav.py', 'name': 'AVAV AeroVironment'},
    {'file': 'get_trading_signal_etn.py', 'name': 'ETN Eaton'},
    {'file': 'get_trading_signal_goog.py', 'name': 'GOOG Google'},
    {'file': 'get_trading_signal_htgc.py', 'name': 'HTGC Hercules Capital Inc'},
    {'file': 'get_trading_signal_nat.py', 'name': 'NAT Nordic American Tankers'},
    {'file': 'get_trading_signal_nxpi.py', 'name': 'NXPI NXP Semiconductors'},
    {'file': 'get_trading_signal_oklo.py', 'name': 'OKLO Oklo Inc'},
    {'file': 'get_trading_signal_omer.py', 'name': 'OMER Omeros Corporation'},
    {'file': 'get_trading_signal_onds.py', 'name': 'ONDS Ondas Holdings'},
    {'file': 'get_trading_signal_rklb.py', 'name': 'RKLB Rocket Lab'},
    {'file': 'get_trading_signal_amd.py', 'name': 'AMD Advanced Micro Devices'},
    {'file': 'get_trading_signal_apld.py', 'name': 'APLD Applied Digital'},
    {'file': 'get_trading_signal_arm.py', 'name': 'ARM Arm Holdings'},
    {'file': 'get_trading_signal_crdo.py', 'name': 'CRDO Credo Technology'},
    {'file': 'get_trading_signal_fn.py', 'name': 'FN Fabrinet'},
    {'file': 'get_trading_signal_gild.py', 'name': 'GILD Gilead Sciences'},
    {'file': 'get_trading_signal_hsai.py', 'name': 'HSAI Hesai Group'},
    {'file': 'get_trading_signal_mrna.py', 'name': 'MRNA Moderna'},
    {'file': 'get_trading_signal_nem.py', 'name': 'NEM Newmont'},
    {'file': 'get_trading_signal_docn.py', 'name': 'DOCN Digital Ocean'},
    {'file': 'get_trading_signal_intc.py', 'name': 'INTC Intel'},
    {'file': 'get_trading_signal_invz.py', 'name': 'INVZ Innoviz Technologies'},
    {'file': 'get_trading_signal_ionq.py', 'name': 'IONQ IonQ'},
    {'file': 'get_trading_signal_klac.py', 'name': 'KLAC KLA Corporation'},
    {'file': 'get_trading_signal_nvo.py', 'name': 'NVO Novo Nordisk'},
    {'file': 'get_trading_signal_oust.py', 'name': 'OUST Ouster'},
    {'file': 'get_trading_signal_qubt.py', 'name': 'QUBT Quantum Computing Inc'},
    {'file': 'get_trading_signal_rdw.py', 'name': 'RDW Redwire'},
    {'file': 'get_trading_signal_rgti.py', 'name': 'RGTI Rigetti Computing'},
    {'file': 'get_trading_signal_satl.py', 'name': 'SATL Satellogic'},
    {'file': 'get_trading_signal_smci.py', 'name': 'SMCI Super Micro Computer'},
    {'file': 'get_trading_signal_smr.py', 'name': 'SMR NuScale Power'},
    {'file': 'get_trading_signal_snow.py', 'name': 'SNOW Snowflake'},
    {'file': 'get_trading_signal_vrt.py', 'name': 'VRT Vertiv Holdings'},
    {'file': 'get_trading_signal_lite.py', 'name': 'LITE Lumentum'},

    # European Stocks
    {'file': 'get_trading_signal_rnmby.py', 'name': 'RNMBY Rheinmetall AG'},
    {'file': 'get_trading_signal_rhm.py', 'name': 'RHM.DE Rheinmetall'},

    # Hong Kong Stocks
    {'file': 'get_trading_signal_02202.py', 'name': '02202.HK Vanke'},
    {'file': 'get_trading_signal_01810.py', 'name': '01810.HK Xiaomi'},

    # Japan Stocks
    {'file': 'get_trading_signal_9984.py', 'name': '9984.T SoftBank Group'},

    # US Stocks - Additional
    {'file': 'get_trading_signal_orcl.py', 'name': 'ORCL Oracle Corporation'},
    {'file': 'get_trading_signal_mdb.py', 'name': 'MDB MongoDB'},
    {'file': 'get_trading_signal_ddog.py', 'name': 'DDOG Datadog'},
    {'file': 'get_trading_signal_stx.py', 'name': 'STX Seagate Technology'},
    {'file': 'get_trading_signal_mpwr.py', 'name': 'MPWR Monolithic Power Systems'},
    {'file': 'get_trading_signal_deck.py', 'name': 'DECK Deckers Outdoor'},
    {'file': 'get_trading_signal_txn.py', 'name': 'TXN Texas Instruments'},
    {'file': 'get_trading_signal_mrvl.py', 'name': 'MRVL Marvell Technology'},
    {'file': 'get_trading_signal_snps.py', 'name': 'SNPS Synopsys'},
    {'file': 'get_trading_signal_cdns.py', 'name': 'CDNS Cadence Design Systems'},
    {'file': 'get_trading_signal_axon.py', 'name': 'AXON Axon Enterprise'},
    {'file': 'get_trading_signal_jazz.py', 'name': 'JAZZ Jazz Pharmaceuticals'},
    {'file': 'get_trading_signal_aaoi.py', 'name': 'AAOI Applied Optoelectronics'},

    # Trained-but-missing entries discovered from model artifacts
    {'file': 'get_trading_signal_amat.py', 'name': 'AMAT AMAT'},
    {'file': 'get_trading_signal_axti.py', 'name': 'AXTI AXT Inc'},
    {'file': 'get_trading_signal_bax.py', 'name': 'BAX BAX'},
    {'file': 'get_trading_signal_cien.py', 'name': 'CIEN CIEN'},
    {'file': 'get_trading_signal_cohr.py', 'name': 'COHR COHR'},
    {'file': 'get_trading_signal_coin.py', 'name': 'COIN COIN'},
    {'file': 'get_trading_signal_glw.py', 'name': 'GLW GLW'},
    {'file': 'get_trading_signal_gpn.py', 'name': 'GPN GPN'},
    {'file': 'get_trading_signal_grmn.py', 'name': 'GRMN GRMN'},
    {'file': 'get_trading_signal_moh.py', 'name': 'MOH MOH'},
    {'file': 'get_trading_signal_omc.py', 'name': 'OMC OMC'},
    {'file': 'get_trading_signal_qcom.py', 'name': 'QCOM QCOM'},
    {'file': 'get_trading_signal_rhm_de.py', 'name': 'RHM.DE RHM.DE'},
    {'file': 'get_trading_signal_tpl.py', 'name': 'TPL TPL'},
    {'file': 'get_trading_signal_viav.py', 'name': 'VIAV VIAV'},
    {'file': 'get_trading_signal_vrtx.py', 'name': 'VRTX VRTX'},

    # New trained stocks
    {'file': 'get_trading_signal_bkr.py',  'name': 'BKR Baker Hughes'},
    {'file': 'get_trading_signal_mchp.py', 'name': 'MCHP Microchip Technology'},
    {'file': 'get_trading_signal_uri.py',  'name': 'URI United Rentals'},
    {'file': 'get_trading_signal_on.py',   'name': 'ON ON Semiconductor'},
    {'file': 'get_trading_signal_gev.py',  'name': 'GEV GE Vernova'},
    {'file': 'get_trading_signal_stld.py', 'name': 'STLD Steel Dynamics'},
    {'file': 'get_trading_signal_ba.py',   'name': 'BA Boeing'},
    {'file': 'get_trading_signal_cost.py', 'name': 'COST Costco'},
    {'file': 'get_trading_signal_spy.py',  'name': 'SPY SP500 ETF'},
    {'file': 'get_trading_signal_asml.py', 'name': 'ASML ASML Holding'},
    {'file': 'get_trading_signal_lrcx.py', 'name': 'LRCX Lam Research'},
    {'file': 'get_trading_signal_simo.py', 'name': 'SIMO Silicon Motion'},
    {'file': 'get_trading_signal_ter.py',  'name': 'TER Teradyne'},

    # Batch 4 — newly trained
    {'file': 'get_trading_signal_jpm.py',   'name': 'JPM JPMorgan Chase'},
    {'file': 'get_trading_signal_xle.py',   'name': 'XLE Energy Select Sector SPDR'},
    {'file': 'get_trading_signal_brk-a.py', 'name': 'BRK-A Berkshire Hathaway A'},
    {'file': 'get_trading_signal_brk-b.py', 'name': 'BRK-B Berkshire Hathaway B'},
    {'file': 'get_trading_signal_mcd.py',   'name': "MCD McDonald's"},
    {'file': 'get_trading_signal_lin.py',   'name': 'LIN Linde'},
    {'file': 'get_trading_signal_rddt.py',  'name': 'RDDT Reddit'},
    {'file': 'get_trading_signal_ko.py',    'name': 'KO Coca-Cola'},
    {'file': 'get_trading_signal_dis.py',   'name': 'DIS Walt Disney'},
    {'file': 'get_trading_signal_crm.py',   'name': 'CRM Salesforce'},
    {'file': 'get_trading_signal_lulu.py',  'name': 'LULU Lululemon Athletica'},
    {'file': 'get_trading_signal_xop.py',   'name': 'XOP SPDR SP Oil Gas E&P ETF'},
    {'file': 'get_trading_signal_smh.py',   'name': 'SMH VanEck Semiconductor ETF'},
    {'file': 'get_trading_signal_shop.py',  'name': 'SHOP Shopify'},
    {'file': 'get_trading_signal_rsp.py',   'name': 'RSP Invesco SP500 Equal Weight ETF'},

    # Batch 5 — newly trained
    {'file': 'get_trading_signal_dxcm.py',  'name': 'DXCM DexCom'},
    {'file': 'get_trading_signal_zs.py',    'name': 'ZS Zscaler'},
    {'file': 'get_trading_signal_ctsh.py',  'name': 'CTSH Cognizant Technology'},
    {'file': 'get_trading_signal_crwd.py',  'name': 'CRWD CrowdStrike'},
    {'file': 'get_trading_signal_rost.py',  'name': 'ROST Ross Stores'},
    {'file': 'get_trading_signal_dell.py',  'name': 'DELL Dell Technologies'},
    {'file': 'get_trading_signal_hpq.py',   'name': 'HPQ HP Inc'},
    {'file': 'get_trading_signal_swks.py',  'name': 'SWKS Skyworks Solutions'},
    {'file': 'get_trading_signal_ntap.py',  'name': 'NTAP NetApp'},
    {'file': 'get_trading_signal_ibm.py',   'name': 'IBM IBM'},
    {'file': 'get_trading_signal_rl.py',    'name': 'RL Ralph Lauren'},
    {'file': 'get_trading_signal_wsm.py',   'name': 'WSM Williams-Sonoma'},
]

def fetch_stock_metrics(ticker_symbol):
    """Fetch sector momentum, beta, volatility, breakout, fundamentals, RSI, MA trend, score"""
    try:
        beta = None
        peg = None
        rev_growth = None
        eps_growth = None
        inst_owned = None
        ma50 = None
        ma200 = None
        # raw (decimal) values used for scoring
        raw_rev = None
        raw_eps = None
        raw_inst = None
        currency = None
        try:
            ticker_obj = yf.Ticker(ticker_symbol)
            info = ticker_obj.info
            currency = info.get('currency') or info.get('financialCurrency')
            beta = info.get('beta', None)
            if beta is not None:
                beta = round(float(beta), 2)
            raw_peg = info.get('pegRatio')
            if raw_peg is not None:
                peg = round(float(raw_peg), 2)
            raw_rev = info.get('revenueGrowth')
            if raw_rev is not None:
                rev_growth = round(float(raw_rev) * 100, 1)
            raw_eps = info.get('earningsGrowth')
            if raw_eps is not None:
                eps_growth = round(float(raw_eps) * 100, 1)
            raw_inst = info.get('heldPercentInstitutions') or info.get('institutionPercentHeld')
            if raw_inst is not None:
                inst_owned = round(float(raw_inst) * 100, 1)
            raw_ma50 = info.get('fiftyDayAverage')
            if raw_ma50 is not None:
                ma50 = round(float(raw_ma50), 2)
            raw_ma200 = info.get('twoHundredDayAverage')
            if raw_ma200 is not None:
                ma200 = round(float(raw_ma200), 2)
        except:
            pass

        df = yf.download(ticker_symbol, period='3mo', progress=False, auto_adjust=True)
        if df.empty or len(df) < 10:
            return (None, beta, None, None, peg, rev_growth, eps_growth,
                    inst_owned, ma50, ma200, None, None, None,
                    None, None, None, None, None, currency, None, None, None)

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)

        close = _as_series(df, 'Close')
        volume = _as_series(df, 'Volume') if 'Volume' in df.columns else None

        latest_price_local = None
        latest_volume = None
        latest_price_usd = None
        turnover_usd = None

        close_clean = close.dropna()
        volume_clean = volume.dropna() if volume is not None else pd.Series(dtype='float64')
        if not close_clean.empty:
            latest_price_local = _safe_float(close_clean.iloc[-1])
        if not volume_clean.empty:
            vol_value = _safe_float(volume_clean.iloc[-1])
            latest_volume = int(vol_value) if vol_value is not None else None

        currency = currency or infer_currency_from_ticker(ticker_symbol)
        fx_rate = get_fx_to_usd(currency)
        if latest_price_local is not None and fx_rate is not None:
            latest_price_usd = round(latest_price_local * fx_rate, 4)
            if latest_volume is not None:
                turnover_usd = latest_price_usd * latest_volume

        # Sector Momentum: 20-day price change %
        sector_momentum = None
        if len(close) >= 20:
            sector_momentum = round(float((close.iloc[-1] - close.iloc[-20]) / close.iloc[-20] * 100), 2)

        # Short-term Volatility: annualized std of 20-day daily returns
        short_vol = None
        returns = close.pct_change().dropna()
        if len(returns) >= 10:
            short_vol = round(float(returns.tail(20).std() * (252 ** 0.5) * 100), 1)

        # Breakout Probability: BB position + volume ratio score
        breakout_prob = None
        if len(df) >= 20:
            bb_mid = close.rolling(20).mean()
            bb_std_val = close.rolling(20).std()
            bb_upper = bb_mid + 2 * bb_std_val
            bb_lower = bb_mid - 2 * bb_std_val
            bb_range = bb_upper.iloc[-1] - bb_lower.iloc[-1]
            if bb_range > 0:
                bb_pos = float((close.iloc[-1] - bb_lower.iloc[-1]) / bb_range * 100)
                vol_mean = volume.rolling(20).mean().iloc[-1] if volume is not None else 0
                vol_ratio = float(volume.iloc[-1] / vol_mean) if volume is not None and vol_mean > 0 else 1.0
                breakout_prob = round(min(100, max(0, bb_pos * 0.5 + min(vol_ratio, 3) / 3 * 50)), 1)

        # RSI (14-period)
        rsi_val = None
        if len(close) >= 15:
            delta = close.diff()
            gain = delta.where(delta > 0, 0).rolling(14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
            rs = gain / (loss + 1e-10)
            rsi_series = 100 - (100 / (1 + rs))
            rsi_val = round(float(rsi_series.iloc[-1]), 1)

        # MA Trend: 1 = bullish (50MA > 200MA), 0 = bearish
        ma_trend = None
        if ma50 is not None and ma200 is not None:
            ma_trend = 1 if ma50 > ma200 else 0

        # Composite score (same formula as user's scoring script)
        score = 0.0
        if peg is not None and peg < 1:
            score += 2
        if raw_rev is not None:
            score += float(raw_rev) * 5
            if raw_rev >= 0.33:
                score += 8  # extra weight for strong revenue growth (>33%)
        if raw_eps is not None:
            score += float(raw_eps) * 3
            if raw_eps >= 1.0:
                score += 8  # extra weight for strong EPS growth (>100%)
        if raw_inst is not None:
            score += float(raw_inst)
        if ma_trend:
            score += 1
        score = round(score, 2)

        # ── MACD 收腳 detection (reuse existing df) ───────────────────────────
        macd_foot = False; macd_gap_up = False; macd_shrink = 0.0
        try:
            if len(close) >= 40:
                ema_f = close.ewm(span=12, adjust=False).mean()
                ema_s = close.ewm(span=26, adjust=False).mean()
                ml = ema_f - ema_s
                ms = ml.ewm(span=9, adjust=False).mean()
                mh = ml - ms
                if len(mh) >= 3:
                    h0, h1, h2 = float(mh.iloc[-1]), float(mh.iloc[-2]), float(mh.iloc[-3])
                    if h0 < 0 and h1 < 0 and h2 < 0 and abs(h0) < abs(h1):
                        shrink = (abs(h1) - abs(h0)) / (abs(h1) + 1e-10) * 100
                        if shrink >= 10:
                            macd_foot = True; macd_shrink = round(shrink, 1)
                            if 'Open' in df.columns and 'High' in df.columns:
                                open_s = _as_series(df, 'Open').dropna()
                                high_s = _as_series(df, 'High').dropna()
                                if len(open_s) >= 1 and len(high_s) >= 2:
                                    macd_gap_up = float(open_s.iloc[-1]) > float(high_s.iloc[-2])
        except Exception:
            pass

        # ── MACD bull status (daily proxy) ────────────────────────────────────
        try:
            ml_last = float((close.ewm(span=12).mean() - close.ewm(span=26).mean()).iloc[-1])
            bull_status = ("✅ 強勢整理" if macd_foot and ml_last > 0
                           else ("🟢 完美多頭" if not macd_foot and ml_last > 0
                                 else "❌ 日線轉負"))
        except Exception:
            bull_status = "—"

        # ── Trailing Stop = Highest Close (20d) − 1.5 × ATR₁₄ ───────────────
        trailing_stop_val = None
        try:
            high_s = _as_series(df, 'High') if 'High' in df.columns else None
            low_s  = _as_series(df, 'Low')  if 'Low'  in df.columns else None
            if high_s is not None and low_s is not None and len(close) >= 15:
                tr = pd.concat([
                    high_s - low_s,
                    (high_s - close.shift(1)).abs(),
                    (low_s  - close.shift(1)).abs(),
                ], axis=1).max(axis=1)
                atr14 = tr.rolling(14).mean()
                highest_close_20 = close.rolling(20).max()
                ts = highest_close_20 - 1.5 * atr14
                trailing_stop_val = round(float(ts.iloc[-1]), 2)
        except Exception:
            pass

        return (sector_momentum, beta, short_vol, breakout_prob, peg, rev_growth,
                eps_growth, inst_owned, ma50, ma200, rsi_val, ma_trend, score,
                macd_foot, macd_gap_up, macd_shrink, bull_status,
                latest_price_local, currency, latest_price_usd, latest_volume,
                turnover_usd, trailing_stop_val)

    except Exception:
        return (None,) * 23

def analyze_volume_and_obv(ticker_symbol):
    """分析资金流入和OBV趋势"""
    try:
        df = yf.download(ticker_symbol, period='3mo', progress=False)

        if df.empty or len(df) < 5:
            return None, None

        # Flatten multi-index columns if necessary
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)

        # 计算OBV
        df['price_change'] = df['Close'].diff()
        df['obv'] = (np.sign(df['price_change']) * df['Volume']).fillna(0).cumsum()

        # 获取最近的数据
        latest = df.iloc[-1]
        prev = df.iloc[-2] if len(df) > 1 else None

        # 检查是否放量收阳
        volume_expansion = False
        positive_close = False

        if prev is not None:
            # 放量：今日成交量 > 前一日成交量
            volume_expansion = latest['Volume'] > prev['Volume']
            # 收阳：收盘价 > 开盘价
            positive_close = latest['Close'] > latest['Open']

        volume_signal = "✅ 放量收阳" if (volume_expansion and positive_close) else ""

        # 检查OBV趋势 (最近5天)
        obv_bullish = False
        if len(df) >= 5:
            obv_recent = df['obv'].iloc[-5:].values
            # 简单判断：OBV整体上升
            obv_bullish = obv_recent[-1] > obv_recent[0]

        obv_signal = "✅ bullish" if obv_bullish else ""

        return volume_signal, obv_signal

    except Exception as e:
        return None, None

def parse_signal_from_output(output):
    """从输出中解析交易信号"""
    output_lower = output.lower()
    if ("🟢 買入信號" in output or "买入信号" in output_lower
            or "(BUY)" in output or "买入 (buy)" in output_lower):
        return "BUY"
    elif ("🔴 賣出信號" in output or "卖出信号" in output_lower
            or "(SELL)" in output or "卖出 (sell)" in output_lower):
        return "SELL"
    elif "(HOLD" in output or "持有 (hold" in output_lower or "强势持有" in output:
        return "HOLD"
    elif "(WAIT)" in output or "观望 (wait)" in output_lower:
        return "WAIT"
    elif "(BUYSELL)" in output or "buysell" in output_lower:
        return "BUYSELL"
    return "UNKNOWN"


def build_summary_rows(results):
    """Fetch Yahoo metrics once, then rank summary rows by USD turnover."""
    rows = []
    total = len(results)

    for pos, result in enumerate(results, 1):
        stock_code, stock_name = split_stock_name(result['name'])
        ticker_for_analysis = normalize_ticker_for_yfinance(stock_code)
        signal = parse_signal_from_output(result['output']) if result['success'] else "N/A"

        row = {
            'result': result,
            'stock_code': stock_code,
            'stock_name': stock_name,
            'ticker_for_analysis': ticker_for_analysis,
            'signal': signal,
            'success': result['success'],
            'volume_signal': None,
            'obv_signal': None,
            'sector_momentum': None,
            'beta': None,
            'short_vol': None,
            'breakout_prob': None,
            'peg': None,
            'rev_growth': None,
            'eps_growth': None,
            'inst_owned': None,
            'ma50': None,
            'ma200': None,
            'rsi_val': None,
            'ma_trend': None,
            'score': None,
            'macd_foot': None,
            'macd_gap_up': None,
            'macd_shrink': None,
            'bull_status': None,
            'latest_price_local': None,
            'currency': None,
            'latest_price_usd': None,
            'latest_volume': None,
            'turnover_usd': None,
            'trailing_stop': None,
        }

        if result['success']:
            print(f"   📊 [{pos}/{total}] 分析 {stock_code} 热度排名、资金流入和OBV趋势...", flush=True)
            time.sleep(1)  # avoid yfinance rate limiting during summary analysis
            row['volume_signal'], row['obv_signal'] = analyze_volume_and_obv(ticker_for_analysis)
            (row['sector_momentum'], row['beta'], row['short_vol'], row['breakout_prob'],
             row['peg'], row['rev_growth'], row['eps_growth'], row['inst_owned'],
             row['ma50'], row['ma200'], row['rsi_val'], row['ma_trend'], row['score'],
             row['macd_foot'], row['macd_gap_up'], row['macd_shrink'], row['bull_status'],
             row['latest_price_local'], row['currency'], row['latest_price_usd'],
             row['latest_volume'], row['turnover_usd'],
             row['trailing_stop']) = fetch_stock_metrics(ticker_for_analysis)

        rows.append(row)

    rows.sort(
        key=lambda r: r['turnover_usd'] if r['turnover_usd'] is not None else -1,
        reverse=True,
    )
    return rows


_ERROR_INDICATORS = ["无法获取数据", "模型加载失败", "信号生成失败", "下载失败", "No data found"]
_SIGNAL_MARKERS  = ["(BUY)", "(SELL)", "(HOLD", "(WAIT)", "(BUYSELL)", "買入信號", "卖出信号"]

async def run_signal_async(script_file, stock_name, semaphore, index, total,
                           max_retries: int = 2, timeout: int = 360):
    """异步运行单个交易信号生成器并捕获输出（自动重试）"""
    async with semaphore:
        script_path = os.path.join(SCRIPT_DIR, script_file)
        last_err = ""

        for attempt in range(1, max_retries + 1):
            prefix = f"\n[{index}/{total}]" if attempt == 1 else f"   ↻ retry {attempt}"
            print(f"{prefix} 运行 {stock_name}...", flush=True)
            try:
                proc = await asyncio.create_subprocess_exec(
                    sys.executable, script_path,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                    cwd=SCRIPT_DIR,
                )
                try:
                    stdout_b, stderr_b = await asyncio.wait_for(
                        proc.communicate(), timeout=timeout)
                except asyncio.TimeoutError:
                    try: proc.kill()
                    except Exception: pass
                    last_err = f"[ERROR] 超时 ({timeout}秒)"
                    print(f"   ⏰ {stock_name} 超时 (attempt {attempt})", flush=True)
                    await asyncio.sleep(random.uniform(5, 12))
                    continue  # retry

                if proc.returncode == 0:
                    stdout = stdout_b.decode('utf-8', errors='ignore')
                    if (any(e in stdout for e in _ERROR_INDICATORS)
                            and not any(s in stdout for s in _SIGNAL_MARKERS)):
                        last_err = f"[ERROR] 无信号\n{stdout[-300:]}"
                        print(f"   ❌ {stock_name} 无信号 (attempt {attempt})", flush=True)
                        await asyncio.sleep(random.uniform(8, 20))
                        continue  # retry
                    print(f"   ✅ {stock_name} 成功", flush=True)
                    return True, stdout
                else:
                    stderr = stderr_b.decode('utf-8', errors='ignore')
                    last_err = f"[ERROR]\n{stderr[-300:]}"
                    print(f"   ❌ {stock_name} 失败 (attempt {attempt})", flush=True)
                    await asyncio.sleep(random.uniform(8, 20))
                    continue  # retry

            except Exception as e:
                last_err = f"[ERROR] {str(e)}"
                print(f"   ❌ {stock_name} 异常: {e} (attempt {attempt})", flush=True)

        return False, last_err

def write_output_to_sheet(ws, stock_name, output_text, success):
    """将输出写入工作表"""
    # 设置标题
    ws['A1'] = stock_name
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4" if success else "C55A11",
                                 end_color="4472C4" if success else "C55A11",
                                 fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')

    # 设置状态
    status = "✅ 成功" if success else "❌ 失败"
    ws['A2'] = "状态"
    ws['B2'] = status
    ws['A2'].font = Font(bold=True)

    # 写入输出内容
    ws['A3'] = "交易信号输出"
    ws['A3'].font = Font(bold=True)
    ws.merge_cells('A3:B3')

    # 将输出按行分割并写入
    lines = output_text.split('\n')
    for idx, line in enumerate(lines, start=4):
        if idx > 1000:  # 限制最多1000行以防止Excel过大
            ws[f'A{idx}'] = "... (输出过长，已截断)"
            break
        ws[f'A{idx}'] = line
        ws.merge_cells(f'A{idx}:B{idx}')

    # 调整列宽
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 50

def create_summary_sheet(wb, results):
    """创建汇总工作表"""
    ws = wb.create_sheet("汇总", 0)  # 插入为第一个工作表

    # 标题
    ws['A1'] = "西方股票交易信号汇总 (US + EU + HK + JP)"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:AA1')

    # 统计信息
    ws['A3'] = "生成时间:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A4'] = "总股票数:"
    ws['B4'] = len(results)
    ws['A5'] = "成功数量:"
    ws['B5'] = sum(1 for r in results if r['success'])
    ws['A6'] = "失败数量:"
    ws['B6'] = sum(1 for r in results if not r['success'])

    for cell in ['A3', 'A4', 'A5', 'A6']:
        ws[cell].font = Font(bold=True)

    summary_rows = build_summary_rows(results)

    # 详细列表标题
    ws['A8'] = "热度排名"
    ws['B8'] = "股票代码"
    ws['C8'] = "股票名称"
    ws['D8'] = "AI信号"
    ws['E8'] = "状态"
    ws['F8'] = "即时资金流入"
    ws['G8'] = "OBV Trend"
    ws['H8'] = "Sector Momentum (20d%)"
    ws['I8'] = "Beta"
    ws['J8'] = "Short-term Volatility"
    ws['K8'] = "Breakout Probability"
    ws['L8'] = "PEG Ratio"
    ws['M8'] = "Revenue Growth"
    ws['N8'] = "EPS Growth"
    ws['O8'] = "Institution Ownership"
    ws['P8'] = "50 Day MA"
    ws['Q8'] = "200 Day MA"
    ws['R8'] = "RSI (14)"
    ws['S8'] = "MA Trend"
    ws['T8'] = "Score"
    ws['U8'] = "MACD收腳"
    ws['V8'] = "跳空缺口"
    ws['W8'] = "MACD格局"
    ws['X8'] = "价格(USD)"
    ws['Y8'] = "成交量"
    ws['Z8'] = "成交额(USD)"
    ws['AA8'] = "原币别"
    ws['AB8'] = "動態止損 (Trail)"

    for cell in ['A8', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8',
                 'L8', 'M8', 'N8', 'O8', 'P8', 'Q8', 'R8', 'S8', 'T8', 'U8', 'V8',
                 'W8', 'X8', 'Y8', 'Z8', 'AA8', 'AB8']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # 写入股票列表
    for idx, row in enumerate(summary_rows, start=9):
        result = row['result']
        stock_code = row['stock_code']
        stock_name = row['stock_name']
        signal = row['signal']
        success = row['success']

        ws[f'A{idx}'] = idx - 8
        ws[f'B{idx}'] = stock_code
        ws[f'C{idx}'] = stock_name

        # 解析AI信号
        ws[f'D{idx}'] = signal

        # 设置AI信号颜色
        if signal == "BUY":
            ws[f'D{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws[f'D{idx}'].font = Font(color="006100", bold=True)
        elif signal == "SELL":
            ws[f'D{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws[f'D{idx}'].font = Font(color="9C0006", bold=True)
        elif signal == "HOLD":
            ws[f'D{idx}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            ws[f'D{idx}'].font = Font(color="2E75B6", bold=True)
        elif signal == "WAIT":
            ws[f'D{idx}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            ws[f'D{idx}'].font = Font(color="9C6500", bold=True)
        elif signal == "BUYSELL":
            ws[f'D{idx}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            ws[f'D{idx}'].font = Font(color="7F4C00", bold=True)
        elif signal == "UNKNOWN":
            ws[f'D{idx}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws[f'D{idx}'].font = Font(color="808080")

        # 状态
        ws[f'E{idx}'] = "✅ 成功" if success else "❌ 失败"

        # 设置状态单元格颜色
        if success:
            ws[f'E{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws[f'E{idx}'].font = Font(color="006100")
        else:
            ws[f'E{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws[f'E{idx}'].font = Font(color="9C0006")

        # 为所有成功的股票分析资金流入和OBV及其他指标
        if success:
            volume_signal = row['volume_signal']
            obv_signal = row['obv_signal']
            sector_momentum = row['sector_momentum']
            beta = row['beta']
            short_vol = row['short_vol']
            breakout_prob = row['breakout_prob']
            peg = row['peg']
            rev_growth = row['rev_growth']
            eps_growth = row['eps_growth']
            inst_owned = row['inst_owned']
            ma50 = row['ma50']
            ma200 = row['ma200']
            rsi_val = row['rsi_val']
            ma_trend = row['ma_trend']
            score = row['score']
            macd_foot = row['macd_foot']
            macd_gap_up = row['macd_gap_up']
            macd_shrink = row['macd_shrink']
            bull_status = row['bull_status']
            latest_price_usd = row['latest_price_usd']
            latest_volume = row['latest_volume']
            turnover_usd = row['turnover_usd']
            currency = row['currency']

            ws[f'F{idx}'] = volume_signal if volume_signal else ""
            ws[f'G{idx}'] = obv_signal if obv_signal else ""

            # Sector Momentum
            if sector_momentum is not None:
                ws[f'H{idx}'] = f"{sector_momentum:+.2f}%"
                if sector_momentum > 5:
                    ws[f'H{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'H{idx}'].font = Font(color="006100")
                elif sector_momentum < -5:
                    ws[f'H{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'H{idx}'].font = Font(color="9C0006")

            # Beta
            if beta is not None:
                ws[f'I{idx}'] = beta
                if beta > 1.5:
                    ws[f'I{idx}'].font = Font(color="9C0006")
                elif beta < 0.5:
                    ws[f'I{idx}'].font = Font(color="006100")

            # Short-term Volatility
            if short_vol is not None:
                ws[f'J{idx}'] = f"{short_vol:.1f}%"
                if short_vol > 50:
                    ws[f'J{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'J{idx}'].font = Font(color="9C0006")

            # Breakout Probability
            if breakout_prob is not None:
                ws[f'K{idx}'] = f"{breakout_prob:.1f}%"
                if breakout_prob >= 70:
                    ws[f'K{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'K{idx}'].font = Font(color="006100", bold=True)
                elif breakout_prob >= 50:
                    ws[f'K{idx}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    ws[f'K{idx}'].font = Font(color="9C6500")

            # PEG Ratio
            if peg is not None:
                ws[f'L{idx}'] = peg
                if peg < 1:
                    ws[f'L{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'L{idx}'].font = Font(color="006100")
                elif peg > 2:
                    ws[f'L{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'L{idx}'].font = Font(color="9C0006")

            # Revenue Growth
            if rev_growth is not None:
                ws[f'M{idx}'] = f"{rev_growth:+.1f}%"
                if rev_growth >= 33:
                    ws[f'M{idx}'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    ws[f'M{idx}'].font = Font(color="1F4E79")
                elif rev_growth > 10:
                    ws[f'M{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'M{idx}'].font = Font(color="006100")
                elif rev_growth < 0:
                    ws[f'M{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'M{idx}'].font = Font(color="9C0006")

            # EPS Growth
            if eps_growth is not None:
                ws[f'N{idx}'] = f"{eps_growth:+.1f}%"
                if eps_growth >= 100:
                    ws[f'N{idx}'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    ws[f'N{idx}'].font = Font(color="1F4E79")
                elif eps_growth > 10:
                    ws[f'N{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'N{idx}'].font = Font(color="006100")
                elif eps_growth < 0:
                    ws[f'N{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'N{idx}'].font = Font(color="9C0006")

            # Institution Ownership
            if inst_owned is not None:
                ws[f'O{idx}'] = f"{inst_owned:.1f}%"

            # 50 Day MA
            if ma50 is not None:
                ws[f'P{idx}'] = ma50

            # 200 Day MA
            if ma200 is not None:
                ws[f'Q{idx}'] = ma200

            # RSI (14)
            if rsi_val is not None:
                ws[f'R{idx}'] = rsi_val
                if rsi_val >= 70:
                    ws[f'R{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'R{idx}'].font = Font(color="9C0006", bold=True)
                elif rsi_val <= 30:
                    ws[f'R{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'R{idx}'].font = Font(color="006100", bold=True)

            # MA Trend
            if ma_trend is not None:
                ws[f'S{idx}'] = "Bullish" if ma_trend == 1 else "Bearish"
                if ma_trend == 1:
                    ws[f'S{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'S{idx}'].font = Font(color="006100")
                else:
                    ws[f'S{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'S{idx}'].font = Font(color="9C0006")

            # Composite Score
            ws[f'T{idx}'] = score if score is not None else ""
            if score is not None and score >= 3:
                ws[f'T{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'T{idx}'].font = Font(color="006100", bold=True)
            elif score is not None and score < 0:
                ws[f'T{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws[f'T{idx}'].font = Font(color="9C0006", bold=True)

            # U: MACD收腳
            if macd_foot:
                ws[f'U{idx}'] = f"✅ {macd_shrink:.0f}%{' ✅跳空' if macd_gap_up else ''}"
                ws[f'U{idx}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                ws[f'U{idx}'].font = Font(color="7F4C00", bold=True)
            else:
                ws[f'U{idx}'] = "—"

            # V: 跳空缺口
            if macd_gap_up:
                ws[f'V{idx}'] = "✅ 跳空"
                ws[f'V{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'V{idx}'].font = Font(color="006100", bold=True)
            else:
                ws[f'V{idx}'] = "—"

            # W: MACD格局
            _colors = {"✅ 強勢整理": ("DDEBF7","2E75B6"),
                       "🟢 完美多頭": ("C6EFCE","006100"),
                       "❌ 日線轉負": ("FFC7CE","9C0006")}
            ws[f'W{idx}'] = bull_status
            if bull_status in _colors:
                bg, fg = _colors[bull_status]
                ws[f'W{idx}'].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                ws[f'W{idx}'].font = Font(color=fg, bold=True)

            # Hotness rank: latest Yahoo close * latest volume, converted to USD.
            if latest_price_usd is not None:
                ws[f'X{idx}'] = latest_price_usd
                ws[f'X{idx}'].number_format = '$#,##0.00'
            if latest_volume is not None:
                ws[f'Y{idx}'] = latest_volume
                ws[f'Y{idx}'].number_format = '#,##0'
            if turnover_usd is not None:
                ws[f'Z{idx}'] = turnover_usd
                ws[f'Z{idx}'].number_format = '$#,##0'
                if idx <= 18:
                    ws[f'Z{idx}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    ws[f'Z{idx}'].font = Font(color="7F4C00", bold=True)
            ws[f'AA{idx}'] = currency if currency else ""

            # AB: Trailing Stop = Highest Close (20d) − 1.5 × ATR₁₄
            trailing_stop = row.get('trailing_stop')
            if trailing_stop is not None:
                ws[f'AB{idx}'] = trailing_stop
                ws[f'AB{idx}'].number_format = '$#,##0.00'
                latest_px = row.get('latest_price_local') or row.get('latest_price_usd')
                if latest_px is not None and latest_px < trailing_stop:
                    # Price already violated trailing stop — red alert
                    ws[f'AB{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'AB{idx}'].font = Font(color="9C0006", bold=True)
                else:
                    ws[f'AB{idx}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    ws[f'AB{idx}'].font = Font(color="375623")
            else:
                ws[f'AB{idx}'] = "—"

            # 如果有信号，设置绿色背景
            if volume_signal:
                ws[f'F{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'F{idx}'].font = Font(color="006100")
            if obv_signal:
                ws[f'G{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'G{idx}'].font = Font(color="006100")
        else:
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                        'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']:
                ws[f'{col}{idx}'] = ""

    ws.freeze_panes = 'A9'
    if summary_rows:
        ws.auto_filter.ref = f"A8:AB{8 + len(summary_rows)}"

    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 22
    ws.column_dimensions['P'].width = 14
    ws.column_dimensions['Q'].width = 14
    ws.column_dimensions['R'].width = 12
    ws.column_dimensions['S'].width = 12
    ws.column_dimensions['T'].width = 10
    ws.column_dimensions['U'].width = 16
    ws.column_dimensions['V'].width = 12
    ws.column_dimensions['W'].width = 14
    ws.column_dimensions['X'].width = 14
    ws.column_dimensions['Y'].width = 14
    ws.column_dimensions['Z'].width = 18
    ws.column_dimensions['AA'].width = 10
    ws.column_dimensions['AB'].width = 18

async def _run_all():
    """Launch all signal scripts concurrently (bounded by CONCURRENT_LIMIT)."""
    semaphore = asyncio.Semaphore(CONCURRENT_LIMIT)
    total = len(SIGNAL_SCRIPTS)
    tasks = [
        run_signal_async(s['file'], s['name'], semaphore, i, total)
        for i, s in enumerate(SIGNAL_SCRIPTS, 1)
    ]
    return await asyncio.gather(*tasks)


if __name__ == "__main__":
    print("=" * 100)
    print(f"批量运行所有西方股票交易信号生成器 (输出到Excel)  [async ×{CONCURRENT_LIMIT}]")
    print("=" * 100)
    print(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"总共 {len(SIGNAL_SCRIPTS)} 个股票  (并行上限: {CONCURRENT_LIMIT})")
    print("=" * 100)

    # ── Run all scripts asynchronously ───────────────────────────────────────
    interrupted = False
    try:
        raw_results = asyncio.run(_run_all())
    except KeyboardInterrupt:
        print("\n\n⚠️  检测到中断 (Ctrl+C)！正在保存已完成的结果...")
        interrupted = True
        raw_results = []   # nothing collected — exit gracefully

    # ── Build Excel workbook from collected results ───────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    results       = []
    success_count = 0
    failed_stocks = []

    for i, (script, (success, output)) in enumerate(zip(SIGNAL_SCRIPTS, raw_results), 1):
        # Sanitise sheet name
        sheet_name = script['name'].translate(str.maketrans('/\\*?:[]', '_______'))
        sheet_name = sheet_name.replace('.', '_')
        if len(sheet_name) > 31:
            sheet_name = script['name'].split()[0].replace('.', '_')

        try:
            ws = wb.create_sheet(sheet_name)
            write_output_to_sheet(ws, script['name'], output, success)
        except Exception as e:
            print(f"   ⚠️  创建工作表失败: {e}")
            ws = wb.create_sheet(f"Stock_{i}")
            write_output_to_sheet(ws, script['name'], output, success)

        results.append({'name': script['name'], 'success': success, 'output': output})

        if success:
            success_count += 1
        else:
            failed_stocks.append(script['name'])

    # ── Summary sheet + save ─────────────────────────────────────────────────
    if results:
        create_summary_sheet(wb, results)

        timestamp       = datetime.now().strftime('%Y%m%d%H%M')
        suffix          = '_PARTIAL' if interrupted else ''
        output_filename = f'western_signals_output_{timestamp}{suffix}.xlsx'
        output_path     = os.path.join(SCRIPT_DIR, output_filename)
        wb.save(output_path)

        print("\n" + "=" * 100)
        print("批量运行被中断! (部分结果已保存)" if interrupted else "批量运行完成!")
        print("=" * 100)
        print(f"成功运行: {success_count}/{len(results)} (总共 {len(SIGNAL_SCRIPTS)} 个)")
        print(f"失败数量: {len(failed_stocks)}")

        if failed_stocks:
            print("\n失败的股票:")
            for stock in failed_stocks:
                print(f"   - {stock}")

        print(f"\n✅ Excel文件已保存: {output_path}")
        if interrupted:
            print(f"⚠️  注意: 这是部分结果 (已完成 {len(results)}/{len(SIGNAL_SCRIPTS)} 个股票)")
        else:
            print("\n所有西方股票信号生成完成!")
    else:
        print("\n❌ 没有生成任何结果")
